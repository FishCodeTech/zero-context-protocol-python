from __future__ import annotations

import asyncio
import json
import math
import time
from dataclasses import asdict, dataclass
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any, Callable, Protocol

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
DEFAULT_EXCEL_REPO = ROOT / "excel-mcp-server"
DEFAULT_OUTPUT_DIR = ROOT / "zero-context-protocol-python" / "benchmark_reports"


@dataclass(frozen=True)
class ScenarioStep:
    tool: str
    arguments: dict[str, Any]
    label: str | None = None


@dataclass(frozen=True)
class ExcelScenario:
    scenario_id: str
    description: str
    steps_factory: Callable[[Path], list[ScenarioStep]]
    verifier: Callable[[Path, dict[str, "CallMetric"]], tuple[bool, str]]


@dataclass
class ToolListMetric:
    tool_names: list[str]
    duration_ms: float
    request_bytes: int
    response_bytes: int


@dataclass
class CallMetric:
    tool: str
    label: str
    ok: bool
    text: str
    raw: dict[str, Any]
    duration_ms: float
    request_bytes: int
    response_bytes: int


@dataclass
class ExcelBenchmarkRun:
    backend_id: str
    client_kind: str
    server_mode: str
    scenario_id: str
    repeat_index: int
    success: bool
    verify_message: str
    total_duration_ms: float
    list_tools_ms: float
    tool_calls: int
    request_bytes: int
    response_bytes: int
    total_bytes: int
    estimated_tokens: int
    step_metrics: list[dict[str, Any]]
    error: str | None = None


@dataclass
class ExcelBenchmarkSummary:
    backend_id: str
    client_kind: str
    server_mode: str
    runs: int
    success_rate: float
    avg_duration_ms: float
    avg_list_tools_ms: float
    avg_request_bytes: float
    avg_response_bytes: float
    avg_total_bytes: float
    avg_estimated_tokens: float


class ExcelBackend(Protocol):
    backend_id: str
    client_kind: str
    server_mode: str

    async def __aenter__(self) -> "ExcelBackend": ...
    async def __aexit__(self, exc_type, exc, tb) -> None: ...
    async def list_tools(self) -> ToolListMetric: ...
    async def call_tool(self, name: str, arguments: dict[str, Any], *, label: str) -> CallMetric: ...


def excel_scenarios() -> list[ExcelScenario]:
    return [
        ExcelScenario(
            scenario_id="workbook_bootstrap_metadata",
            description="Create a workbook, add sheets, and verify workbook metadata.",
            steps_factory=_bootstrap_steps,
            verifier=_verify_bootstrap,
        ),
        ExcelScenario(
            scenario_id="sales_table_and_formulas",
            description="Write tabular sales data, apply formulas, format the header, and create an Excel table.",
            steps_factory=_sales_steps,
            verifier=_verify_sales,
        ),
        ExcelScenario(
            scenario_id="report_layout_and_readback",
            description="Build a simple report sheet, merge the title row, and read the final range back.",
            steps_factory=_report_steps,
            verifier=_verify_report,
        ),
    ]


def backend_factories(
    *,
    python_executable: str,
    excel_repo: Path,
) -> list[Callable[[], ExcelBackend]]:
    return [
        lambda: MCPProcessBackend(
            backend_id="mcp_client_to_original_mcp",
            client_kind="mcp_client",
            server_mode="original_mcp",
            python_executable=python_executable,
            excel_repo=excel_repo,
            server_args=["-m", "excel_mcp", "stdio"],
        ),
        lambda: ZCPProxyBackend(
            backend_id="zcp_client_to_original_mcp",
            client_kind="zcp_client",
            server_mode="original_mcp",
            python_executable=python_executable,
            excel_repo=excel_repo,
            server_args=["-m", "excel_mcp", "stdio"],
        ),
        lambda: NativeZCPBackend(),
        lambda: MCPProcessBackend(
            backend_id="mcp_client_to_zcp_mcp_surface",
            client_kind="mcp_client",
            server_mode="native_zcp",
            python_executable=python_executable,
            excel_repo=excel_repo,
            server_args=["-m", "excel_mcp", "zcp-mcp-stdio"],
        ),
    ]


async def run_excel_client_benchmark_async(
    *,
    repeats: int = 3,
    python_executable: str,
    excel_repo: Path = DEFAULT_EXCEL_REPO,
) -> dict[str, Any]:
    scenarios = excel_scenarios()
    runs: list[ExcelBenchmarkRun] = []

    for factory in backend_factories(python_executable=python_executable, excel_repo=excel_repo):
        for repeat_index in range(repeats):
            async with factory() as backend:
                tool_listing = await backend.list_tools()
                available = set(tool_listing.tool_names)
                for scenario in scenarios:
                    steps = scenario.steps_factory(excel_repo)
                    missing = [step.tool for step in steps if step.tool not in available]
                    if missing:
                        runs.append(
                            ExcelBenchmarkRun(
                                backend_id=backend.backend_id,
                                client_kind=backend.client_kind,
                                server_mode=backend.server_mode,
                                scenario_id=scenario.scenario_id,
                                repeat_index=repeat_index,
                                success=False,
                                verify_message=f"missing tools: {', '.join(sorted(set(missing)))}",
                                total_duration_ms=0.0,
                                list_tools_ms=tool_listing.duration_ms,
                                tool_calls=0,
                                request_bytes=tool_listing.request_bytes,
                                response_bytes=tool_listing.response_bytes,
                                total_bytes=tool_listing.request_bytes + tool_listing.response_bytes,
                                estimated_tokens=_estimate_tokens(tool_listing.request_bytes + tool_listing.response_bytes),
                                step_metrics=[],
                                error="missing_tools",
                            )
                        )
                        continue

                    with TemporaryDirectory(prefix=f"{backend.backend_id}-{scenario.scenario_id}-") as temp_dir:
                        start = time.perf_counter()
                        step_results: dict[str, CallMetric] = {}
                        request_bytes = tool_listing.request_bytes
                        response_bytes = tool_listing.response_bytes
                        error: str | None = None
                        try:
                            for step_index, step in enumerate(steps):
                                label = step.label or f"{step.tool}_{step_index + 1}"
                                prepared = _resolve_arguments(step.arguments, Path(temp_dir))
                                metric = await backend.call_tool(step.tool, prepared, label=label)
                                step_results[label] = metric
                                request_bytes += metric.request_bytes
                                response_bytes += metric.response_bytes
                                if not metric.ok:
                                    raise RuntimeError(f"{step.tool} failed: {metric.text}")
                            success, verify_message = scenario.verifier(Path(temp_dir), step_results)
                        except Exception as exc:
                            success = False
                            verify_message = str(exc)
                            error = str(exc)
                        total_duration_ms = (time.perf_counter() - start) * 1000
                        total_bytes = request_bytes + response_bytes
                        runs.append(
                            ExcelBenchmarkRun(
                                backend_id=backend.backend_id,
                                client_kind=backend.client_kind,
                                server_mode=backend.server_mode,
                                scenario_id=scenario.scenario_id,
                                repeat_index=repeat_index,
                                success=success,
                                verify_message=verify_message,
                                total_duration_ms=total_duration_ms,
                                list_tools_ms=tool_listing.duration_ms,
                                tool_calls=len(step_results),
                                request_bytes=request_bytes,
                                response_bytes=response_bytes,
                                total_bytes=total_bytes,
                                estimated_tokens=_estimate_tokens(total_bytes),
                                step_metrics=[asdict(metric) for metric in step_results.values()],
                                error=error,
                            )
                        )

    summaries = summarize_excel_runs(runs)
    return {
        "repeats": repeats,
        "excel_repo": str(excel_repo),
        "runs": [asdict(run) for run in runs],
        "summary": [asdict(summary) for summary in summaries],
    }


def run_excel_client_benchmark(
    *,
    repeats: int = 3,
    python_executable: str,
    excel_repo: Path = DEFAULT_EXCEL_REPO,
) -> dict[str, Any]:
    return asyncio.run(
        run_excel_client_benchmark_async(
            repeats=repeats,
            python_executable=python_executable,
            excel_repo=excel_repo,
        )
    )


def summarize_excel_runs(runs: list[ExcelBenchmarkRun]) -> list[ExcelBenchmarkSummary]:
    buckets: dict[str, list[ExcelBenchmarkRun]] = {}
    for run in runs:
        buckets.setdefault(run.backend_id, []).append(run)

    summaries: list[ExcelBenchmarkSummary] = []
    for backend_id, backend_runs in buckets.items():
        first = backend_runs[0]
        total = len(backend_runs)
        summaries.append(
            ExcelBenchmarkSummary(
                backend_id=backend_id,
                client_kind=first.client_kind,
                server_mode=first.server_mode,
                runs=total,
                success_rate=sum(1 for run in backend_runs if run.success) / total,
                avg_duration_ms=sum(run.total_duration_ms for run in backend_runs) / total,
                avg_list_tools_ms=sum(run.list_tools_ms for run in backend_runs) / total,
                avg_request_bytes=sum(run.request_bytes for run in backend_runs) / total,
                avg_response_bytes=sum(run.response_bytes for run in backend_runs) / total,
                avg_total_bytes=sum(run.total_bytes for run in backend_runs) / total,
                avg_estimated_tokens=sum(run.estimated_tokens for run in backend_runs) / total,
            )
        )
    summaries.sort(key=lambda item: (item.server_mode, item.client_kind, item.backend_id))
    return summaries


def markdown_excel_report(report: dict[str, Any]) -> str:
    lines = [
        "# Excel Client Protocol Benchmark",
        "",
        f"- repeats: `{report['repeats']}`",
        f"- excel repo: `{report['excel_repo']}`",
        "",
        "## Summary",
        "",
        "| Backend | Client | Server Mode | Success | Avg Duration (ms) | Avg Payload Bytes | Avg Token Estimate |",
        "| --- | --- | --- | --- | ---: | ---: | ---: |",
    ]
    for summary in report["summary"]:
        lines.append(
            "| {backend_id} | {client_kind} | {server_mode} | {success_rate:.1%} | {avg_duration_ms:.1f} | {avg_total_bytes:.1f} | {avg_estimated_tokens:.1f} |".format(
                **summary
            )
        )

    lines.extend(["", "## Scenario Runs", ""])
    by_scenario: dict[str, list[dict[str, Any]]] = {}
    for run in report["runs"]:
        by_scenario.setdefault(run["scenario_id"], []).append(run)

    for scenario_id, scenario_runs in sorted(by_scenario.items()):
        lines.extend(
            [
                f"### `{scenario_id}`",
                "",
                "| Backend | Success | Duration (ms) | Tool Calls | Payload Bytes | Verify |",
                "| --- | --- | ---: | ---: | ---: | --- |",
            ]
        )
        for run in scenario_runs:
            lines.append(
                "| {backend_id} | {success} | {total_duration_ms:.1f} | {tool_calls} | {total_bytes} | {verify_message} |".format(
                    **run
                )
            )
        lines.append("")

    return "\n".join(lines).strip() + "\n"


def write_excel_reports(report: dict[str, Any], output_dir: Path = DEFAULT_OUTPUT_DIR) -> tuple[Path, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    json_path = output_dir / "excel_client_protocol_benchmark.json"
    markdown_path = output_dir / "excel_client_protocol_benchmark.md"
    json_path.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    markdown_path.write_text(markdown_excel_report(report), encoding="utf-8")
    return json_path, markdown_path


class MCPProcessBackend:
    def __init__(
        self,
        *,
        backend_id: str,
        client_kind: str,
        server_mode: str,
        python_executable: str,
        excel_repo: Path,
        server_args: list[str],
    ) -> None:
        self.backend_id = backend_id
        self.client_kind = client_kind
        self.server_mode = server_mode
        self.python_executable = python_executable
        self.excel_repo = excel_repo
        self.server_args = server_args
        self._request_id = 0
        self._stdio_cm = None
        self._session_cm = None
        self._session = None

    async def __aenter__(self) -> "MCPProcessBackend":
        from mcp.client.session import ClientSession
        from mcp.client.stdio import StdioServerParameters, stdio_client

        self._stdio_cm = stdio_client(
            StdioServerParameters(
                command=self.python_executable,
                args=self.server_args,
                cwd=str(self.excel_repo),
            )
        )
        read_stream, write_stream = await self._stdio_cm.__aenter__()
        self._session_cm = ClientSession(read_stream, write_stream)
        self._session = await self._session_cm.__aenter__()
        await self._session.initialize()
        return self

    async def __aexit__(self, exc_type, exc, tb) -> None:
        if self._session_cm is not None:
            await self._session_cm.__aexit__(exc_type, exc, tb)
        if self._stdio_cm is not None:
            await self._stdio_cm.__aexit__(exc_type, exc, tb)

    async def list_tools(self) -> ToolListMetric:
        self._request_id += 1
        request_bytes = _json_size({"jsonrpc": "2.0", "id": self._request_id, "method": "tools/list", "params": {}})
        start = time.perf_counter()
        result = await self._session.list_tools()
        duration_ms = (time.perf_counter() - start) * 1000
        payload = {
            "tools": [tool.model_dump(by_alias=True, exclude_none=True) for tool in result.tools],
            "nextCursor": getattr(result, "nextCursor", None),
        }
        response_bytes = _json_size(payload)
        return ToolListMetric(
            tool_names=[tool.name for tool in result.tools],
            duration_ms=duration_ms,
            request_bytes=request_bytes,
            response_bytes=response_bytes,
        )

    async def call_tool(self, name: str, arguments: dict[str, Any], *, label: str) -> CallMetric:
        self._request_id += 1
        request_bytes = _json_size(
            {
                "jsonrpc": "2.0",
                "id": self._request_id,
                "method": "tools/call",
                "params": {"name": name, "arguments": arguments},
            }
        )
        start = time.perf_counter()
        result = await self._session.call_tool(name, arguments)
        duration_ms = (time.perf_counter() - start) * 1000
        payload = result.model_dump(exclude_none=True, by_alias=True)
        text = _extract_tool_text(payload)
        return CallMetric(
            tool=name,
            label=label,
            ok=not payload.get("isError", False) and not text.startswith("Error:"),
            text=text,
            raw=payload,
            duration_ms=duration_ms,
            request_bytes=request_bytes,
            response_bytes=_json_size(payload),
        )


class ZCPProxyBackend:
    def __init__(
        self,
        *,
        backend_id: str,
        client_kind: str,
        server_mode: str,
        python_executable: str,
        excel_repo: Path,
        server_args: list[str],
    ) -> None:
        self.backend_id = backend_id
        self.client_kind = client_kind
        self.server_mode = server_mode
        self.python_executable = python_executable
        self.excel_repo = excel_repo
        self.server_args = server_args
        self._request_id = 0
        self._stdio_cm = None
        self._session_cm = None
        self._mcp_session = None
        self._zcp_client = None

    async def __aenter__(self) -> "ZCPProxyBackend":
        from mcp.client.session import ClientSession
        from mcp.client.stdio import StdioServerParameters, stdio_client as mcp_stdio_client
        from zcp import ZCPClientSession

        self._stdio_cm = mcp_stdio_client(
            StdioServerParameters(
                command=self.python_executable,
                args=self.server_args,
                cwd=str(self.excel_repo),
            )
        )
        read_stream, write_stream = await self._stdio_cm.__aenter__()
        self._session_cm = ClientSession(read_stream, write_stream)
        self._mcp_session = await self._session_cm.__aenter__()
        await self._mcp_session.initialize()

        relay = MCPRelayServerSession(self._mcp_session)
        self._zcp_client = ZCPClientSession(relay, transport="mcp-relay")
        await self._zcp_client.initialize()
        await self._zcp_client.initialized()
        return self

    async def __aexit__(self, exc_type, exc, tb) -> None:
        if self._session_cm is not None:
            await self._session_cm.__aexit__(exc_type, exc, tb)
        if self._stdio_cm is not None:
            await self._stdio_cm.__aexit__(exc_type, exc, tb)

    async def list_tools(self) -> ToolListMetric:
        self._request_id += 1
        request_bytes = _json_size({"jsonrpc": "2.0", "id": self._request_id, "method": "tools/list", "params": {}})
        start = time.perf_counter()
        result = await self._zcp_client.list_tools()
        duration_ms = (time.perf_counter() - start) * 1000
        return ToolListMetric(
            tool_names=[tool["name"] for tool in result["tools"]],
            duration_ms=duration_ms,
            request_bytes=request_bytes,
            response_bytes=_json_size(result),
        )

    async def call_tool(self, name: str, arguments: dict[str, Any], *, label: str) -> CallMetric:
        self._request_id += 1
        request_bytes = _json_size(
            {
                "jsonrpc": "2.0",
                "id": self._request_id,
                "method": "tools/call",
                "params": {"name": name, "arguments": arguments},
            }
        )
        start = time.perf_counter()
        result = await self._zcp_client.call_tool(name, arguments)
        duration_ms = (time.perf_counter() - start) * 1000
        text = _extract_tool_text(result)
        return CallMetric(
            tool=name,
            label=label,
            ok=not result.get("isError", False) and not text.startswith("Error:"),
            text=text,
            raw=result,
            duration_ms=duration_ms,
            request_bytes=request_bytes,
            response_bytes=_json_size(result),
        )


class NativeZCPBackend:
    backend_id = "zcp_client_to_native_zcp"
    client_kind = "zcp_client"
    server_mode = "native_zcp"

    def __init__(self) -> None:
        self._request_id = 0
        self._client = None

    async def __aenter__(self) -> "NativeZCPBackend":
        from excel_mcp.zcp_server import build_excel_zcp_app
        from zcp import stdio_client, stdio_server

        app = build_excel_zcp_app()
        self._client = stdio_client(stdio_server(app))
        await self._client.initialize()
        await self._client.initialized()
        return self

    async def __aexit__(self, exc_type, exc, tb) -> None:
        return None

    async def list_tools(self) -> ToolListMetric:
        self._request_id += 1
        request_bytes = _json_size({"jsonrpc": "2.0", "id": self._request_id, "method": "tools/list", "params": {}})
        start = time.perf_counter()
        result = await self._client.list_tools()
        duration_ms = (time.perf_counter() - start) * 1000
        return ToolListMetric(
            tool_names=[tool["name"] for tool in result["tools"]],
            duration_ms=duration_ms,
            request_bytes=request_bytes,
            response_bytes=_json_size(result),
        )

    async def call_tool(self, name: str, arguments: dict[str, Any], *, label: str) -> CallMetric:
        self._request_id += 1
        request_bytes = _json_size(
            {
                "jsonrpc": "2.0",
                "id": self._request_id,
                "method": "tools/call",
                "params": {"name": name, "arguments": arguments},
            }
        )
        start = time.perf_counter()
        result = await self._client.call_tool(name, arguments)
        duration_ms = (time.perf_counter() - start) * 1000
        text = _extract_tool_text(result)
        return CallMetric(
            tool=name,
            label=label,
            ok=not result.get("isError", False) and not text.startswith("Error:"),
            text=text,
            raw=result,
            duration_ms=duration_ms,
            request_bytes=request_bytes,
            response_bytes=_json_size(result),
        )


def _bootstrap_steps(_repo_root: Path) -> list[ScenarioStep]:
    return [
        ScenarioStep("create_workbook", {"filepath": "{workdir}/bootstrap.xlsx"}),
        ScenarioStep("create_worksheet", {"filepath": "{workdir}/bootstrap.xlsx", "sheet_name": "RawData"}),
        ScenarioStep("create_worksheet", {"filepath": "{workdir}/bootstrap.xlsx", "sheet_name": "Summary"}),
        ScenarioStep(
            "get_workbook_metadata",
            {"filepath": "{workdir}/bootstrap.xlsx", "include_ranges": True},
            label="metadata",
        ),
    ]


def _sales_steps(_repo_root: Path) -> list[ScenarioStep]:
    return [
        ScenarioStep("create_workbook", {"filepath": "{workdir}/sales.xlsx"}),
        ScenarioStep("create_worksheet", {"filepath": "{workdir}/sales.xlsx", "sheet_name": "Sales"}),
        ScenarioStep(
            "write_data_to_excel",
            {
                "filepath": "{workdir}/sales.xlsx",
                "sheet_name": "Sales",
                "data": [
                    ["Region", "Q1", "Q2", "Total"],
                    ["East", 120, 140, None],
                    ["West", 90, 110, None],
                    ["Central", 75, 88, None],
                ],
                "start_cell": "A1",
            },
        ),
        ScenarioStep(
            "validate_formula_syntax",
            {"filepath": "{workdir}/sales.xlsx", "sheet_name": "Sales", "cell": "D2", "formula": "=SUM(B2:C2)"},
        ),
        ScenarioStep("apply_formula", {"filepath": "{workdir}/sales.xlsx", "sheet_name": "Sales", "cell": "D2", "formula": "=SUM(B2:C2)"}),
        ScenarioStep("apply_formula", {"filepath": "{workdir}/sales.xlsx", "sheet_name": "Sales", "cell": "D3", "formula": "=SUM(B3:C3)"}),
        ScenarioStep("apply_formula", {"filepath": "{workdir}/sales.xlsx", "sheet_name": "Sales", "cell": "D4", "formula": "=SUM(B4:C4)"}),
        ScenarioStep(
            "format_range",
            {
                "filepath": "{workdir}/sales.xlsx",
                "sheet_name": "Sales",
                "start_cell": "A1",
                "end_cell": "D1",
                "bold": True,
                "font_color": "FFFFFF",
                "bg_color": "1F4E78",
                "alignment": "center",
                "wrap_text": True,
            },
        ),
        ScenarioStep(
            "create_table",
            {
                "filepath": "{workdir}/sales.xlsx",
                "sheet_name": "Sales",
                "data_range": "A1:D4",
                "table_name": "SalesTable",
                "table_style": "TableStyleMedium9",
            },
            label="table",
        ),
    ]


def _report_steps(_repo_root: Path) -> list[ScenarioStep]:
    return [
        ScenarioStep("create_workbook", {"filepath": "{workdir}/report.xlsx"}),
        ScenarioStep("create_worksheet", {"filepath": "{workdir}/report.xlsx", "sheet_name": "Report"}),
        ScenarioStep(
            "write_data_to_excel",
            {
                "filepath": "{workdir}/report.xlsx",
                "sheet_name": "Report",
                "data": [
                    ["Quarterly Sales Report", "", "", ""],
                    ["Region", "Q1", "Q2", "Total"],
                    ["East", 120, 140, 260],
                    ["West", 90, 110, 200],
                ],
                "start_cell": "A1",
            },
        ),
        ScenarioStep("merge_cells", {"filepath": "{workdir}/report.xlsx", "sheet_name": "Report", "start_cell": "A1", "end_cell": "D1"}),
        ScenarioStep(
            "format_range",
            {
                "filepath": "{workdir}/report.xlsx",
                "sheet_name": "Report",
                "start_cell": "A1",
                "end_cell": "D1",
                "bold": True,
                "alignment": "center",
                "bg_color": "D9EAF7",
            },
        ),
        ScenarioStep("get_merged_cells", {"filepath": "{workdir}/report.xlsx", "sheet_name": "Report"}, label="merged"),
        ScenarioStep(
            "read_data_from_excel",
            {
                "filepath": "{workdir}/report.xlsx",
                "sheet_name": "Report",
                "start_cell": "A1",
                "end_cell": "D4",
                "preview_only": False,
            },
            label="readback",
        ),
    ]


def _verify_bootstrap(workdir: Path, step_results: dict[str, CallMetric]) -> tuple[bool, str]:
    workbook_path = workdir / "bootstrap.xlsx"
    if not workbook_path.exists():
        return False, "bootstrap.xlsx was not created"
    workbook = load_workbook(workbook_path)
    try:
        sheet_names = set(workbook.sheetnames)
    finally:
        workbook.close()
    ok = {"RawData", "Summary"}.issubset(sheet_names)
    metadata_text = step_results["metadata"].text
    return ok and "RawData" in metadata_text and "Summary" in metadata_text, f"sheets={sorted(sheet_names)}"


def _verify_sales(workdir: Path, step_results: dict[str, CallMetric]) -> tuple[bool, str]:
    workbook_path = workdir / "sales.xlsx"
    workbook = load_workbook(workbook_path)
    try:
        worksheet = workbook["Sales"]
        formulas_ok = (
            worksheet["D2"].value == "=SUM(B2:C2)"
            and worksheet["D3"].value == "=SUM(B3:C3)"
            and worksheet["D4"].value == "=SUM(B4:C4)"
        )
        table_ok = "SalesTable" in worksheet.tables
        header_ok = bool(worksheet["A1"].font.bold)
    finally:
        workbook.close()
    ok = formulas_ok and table_ok and header_ok
    return ok, f"formulas={formulas_ok}, table={table_ok}, header_bold={header_ok}, table_result={step_results['table'].text}"


def _verify_report(workdir: Path, step_results: dict[str, CallMetric]) -> tuple[bool, str]:
    workbook_path = workdir / "report.xlsx"
    workbook = load_workbook(workbook_path)
    try:
        worksheet = workbook["Report"]
        merged_ranges = {str(item) for item in worksheet.merged_cells.ranges}
        title = worksheet["A1"].value
    finally:
        workbook.close()
    readback = _parse_json_object(step_results["readback"].text)
    merged_ok = "A1:D1" in merged_ranges
    cells = readback.get("cells", []) if isinstance(readback, dict) else []
    title_ok = title == "Quarterly Sales Report"
    readback_ok = any(cell.get("address") == "A1" and cell.get("value") == "Quarterly Sales Report" for cell in cells)
    ok = merged_ok and title_ok and readback_ok
    return ok, f"merged={sorted(merged_ranges)}, merged_tool={step_results['merged'].text}"


def _resolve_arguments(arguments: dict[str, Any], workdir: Path) -> dict[str, Any]:
    def resolve(value: Any) -> Any:
        if isinstance(value, str):
            return value.replace("{workdir}", str(workdir))
        if isinstance(value, list):
            return [resolve(item) for item in value]
        if isinstance(value, dict):
            return {key: resolve(item) for key, item in value.items()}
        return value

    return {key: resolve(value) for key, value in arguments.items()}


def _json_size(payload: dict[str, Any]) -> int:
    return len(json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8"))


def _estimate_tokens(total_bytes: int) -> int:
    return math.ceil(total_bytes / 4)


def _extract_tool_text(payload: dict[str, Any]) -> str:
    structured = payload.get("structuredContent")
    if isinstance(structured, dict) and set(structured) == {"result"} and isinstance(structured["result"], str):
        return structured["result"]
    if isinstance(structured, str):
        return structured
    if isinstance(structured, (dict, list)):
        return json.dumps(structured, ensure_ascii=False)

    content = payload.get("content")
    if isinstance(content, list):
        parts = [item.get("text", "") for item in content if isinstance(item, dict)]
        return "".join(parts)
    if isinstance(content, str):
        return content

    if payload.get("error"):
        return str(payload["error"])
    return json.dumps(payload, ensure_ascii=False)


def _parse_json_object(text: str) -> dict[str, Any]:
    try:
        parsed = json.loads(text)
        if isinstance(parsed, str):
            parsed = json.loads(parsed)
        return parsed if isinstance(parsed, dict) else {}
    except Exception:
        return {}


class MCPRelayServerSession:
    def __init__(self, mcp_session: Any) -> None:
        self.mcp_session = mcp_session
        self._notifications: list[dict[str, Any]] = []

    def attach_client(self, **_: Any) -> None:
        return None

    def drain_notifications(self) -> list[dict[str, Any]]:
        pending = list(self._notifications)
        self._notifications.clear()
        return pending

    async def handle_message(self, message: dict[str, Any]) -> dict[str, Any]:
        request_id = message.get("id")
        method = message["method"]
        if method == "initialize":
            return {
                "jsonrpc": "2.0",
                "id": request_id,
                "result": {
                    "protocolVersion": "2026-03-01",
                    "serverInfo": {"name": "excel-mcp-relay", "version": "0.1.0"},
                    "capabilities": {"tools": {"listChanged": False}},
                },
            }
        if method == "initialized":
            return {"jsonrpc": "2.0", "id": request_id, "result": {"ok": True}}
        if method == "tools/list":
            listed = await self.mcp_session.list_tools()
            return {
                "jsonrpc": "2.0",
                "id": request_id,
                "result": {
                    "tools": [
                        {
                            "name": tool.name,
                            "description": tool.description or tool.title or tool.name,
                            "inputSchema": tool.inputSchema,
                            "outputSchema": tool.outputSchema,
                            "annotations": tool.annotations.model_dump(by_alias=True, exclude_none=True) if tool.annotations else None,
                        }
                        for tool in listed.tools
                    ]
                },
            }
        if method == "tools/call":
            result = await self.mcp_session.call_tool(message["params"]["name"], message["params"].get("arguments", {}))
            dumped = result.model_dump(exclude_none=True, by_alias=True)
            text = _extract_tool_text(dumped)
            if dumped.get("isError") or text.startswith("Error:"):
                return {"jsonrpc": "2.0", "id": request_id, "result": {"isError": True, "error": text}}
            return {
                "jsonrpc": "2.0",
                "id": request_id,
                "result": {
                    "isError": False,
                    "content": [{"type": "text", "text": text}],
                    "structuredContent": text,
                    "summary": f"{message['params']['name']}: {text[:64]}",
                },
            }
        return {"jsonrpc": "2.0", "id": request_id, "error": {"code": -32601, "message": f"unsupported:{method}"}}
