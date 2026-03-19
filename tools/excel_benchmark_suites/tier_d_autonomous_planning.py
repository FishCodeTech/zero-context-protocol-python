from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable

from openpyxl import Workbook, load_workbook

from .common import ExcelBenchmarkCase


@dataclass(frozen=True)
class SuiteCase:
    case_id: str
    prompt_factory: Callable[[Path], str]
    expected_tool_calls: dict[str, int]
    evaluator: Callable[[dict[str, Any] | None, Path], tuple[bool, bool, str]]


def _load_workbook_safely(workbook: Path) -> tuple[Workbook | None, str | None]:
    if not workbook.exists():
        return None, "workbook_missing"
    try:
        return load_workbook(workbook), None
    except Exception as exc:
        return None, f"workbook_load_error:{exc}"


def _sheet_or_note(wb: Workbook | None, sheet_name: str) -> tuple[Any | None, str | None]:
    if wb is None:
        return None, "workbook_unavailable"
    if sheet_name not in wb.sheetnames:
        return None, f"missing_sheet:{sheet_name};available={wb.sheetnames}"
    return wb[sheet_name], None


def tier_d_autonomous_planning_cases(case_cls: type | None = None) -> list[Any]:
    cls = case_cls or SuiteCase
    specs = [
        (
            "executive_briefing_goal",
            _executive_briefing_prompt,
            {
                "create_workbook": 1,
                "create_worksheet": 1,
                "write_data_to_excel": 2,
                "merge_cells": 1,
                "format_range": 1,
                "get_merged_cells": 1,
                "read_data_from_excel": 1,
            },
            _evaluate_executive_briefing,
        ),
        (
            "team_plan_snapshot_goal",
            _team_plan_snapshot_prompt,
            {
                "create_workbook": 1,
                "create_worksheet": 1,
                "write_data_to_excel": 1,
                "copy_worksheet": 1,
                "rename_worksheet": 1,
                "get_workbook_metadata": 1,
            },
            _evaluate_team_plan_snapshot,
        ),
        (
            "month_end_close_goal",
            _month_end_close_prompt,
            {
                "create_workbook": 1,
                "create_worksheet": 2,
                "write_data_to_excel": 3,
                "apply_formula": 2,
                "read_data_from_excel": 1,
            },
            _evaluate_month_end_close,
        ),
        (
            "requests_register_goal",
            _requests_register_prompt,
            {
                "create_workbook": 1,
                "create_worksheet": 1,
                "write_data_to_excel": 1,
                "create_table": 1,
                "format_range": 1,
                "read_data_from_excel": 1,
            },
            _evaluate_requests_register,
        ),
        (
            "board_packet_repair_goal",
            _board_packet_repair_prompt,
            {
                "rename_worksheet": 1,
                "create_worksheet": 1,
                "write_data_to_excel": 1,
                "merge_cells": 1,
                "format_range": 1,
                "get_workbook_metadata": 1,
            },
            _evaluate_board_packet_repair,
        ),
        (
            "staging_cleanup_goal",
            _staging_cleanup_prompt,
            {
                "delete_sheet_rows": 1,
                "delete_sheet_columns": 1,
                "read_data_from_excel": 1,
            },
            _evaluate_staging_cleanup,
        ),
    ]
    return [
        cls(
            case_id=case_id,
            prompt_factory=prompt_factory,
            expected_tool_calls=expected_tool_calls,
            evaluator=evaluator,
        )
        for case_id, prompt_factory, expected_tool_calls, evaluator in specs
    ]


def tier_d_cases() -> list[ExcelBenchmarkCase]:
    semantic_tools = {
        "executive_briefing_goal": {"zcp_finalize_executive_briefing": 1},
        "team_plan_snapshot_goal": {"zcp_create_team_plan_snapshot": 1},
        "month_end_close_goal": {"zcp_finalize_month_end_close": 1},
        "requests_register_goal": {"zcp_build_requests_register": 1},
        "board_packet_repair_goal": {"zcp_repair_board_packet": 1},
        "staging_cleanup_goal": {"zcp_cleanup_staging_backlog": 1},
    }
    return [
        ExcelBenchmarkCase(
            tier="D",
            case_id=case.case_id,
            prompt_factory=case.prompt_factory,
            required_tool_calls=case.expected_tool_calls,
            evaluator=case.evaluator,
            autonomous=True,
            native_zcp_required_tool_calls=semantic_tools[case.case_id],
        )
        for case in tier_d_autonomous_planning_cases()
    ]


def _executive_briefing_prompt(workdir: Path) -> str:
    workbook = workdir / "executive_briefing.xlsx"
    return (
        f"请为运营周会准备一个新的 Excel 简报文件，保存到绝对路径 `{workbook}`。"
        " 目标是让管理层打开文件后马上看到一个清晰的摘要页：顶部要有一个跨列标题区，下面有三项核心指标"
        " `Revenue`、`Cost`、`Profit` 以及对应数字。"
        " 你可以自行规划具体步骤，但最终文件必须可读、结构清楚，并在结束前自行检查结果。"
        ' 最终只输出一行 JSON，格式为 {"workbook":"executive_briefing.xlsx","sheet":"Dashboard","merged_ranges":["A1:D1"],"metrics":["Revenue","Cost","Profit"]}。'
    )


def _team_plan_snapshot_prompt(workdir: Path) -> str:
    workbook = workdir / "team_plan_snapshot.xlsx"
    return (
        f"请创建一个团队排班计划 Excel，保存到绝对路径 `{workbook}`。"
        " 业务目标是既要有当前可编辑的计划页，也要有一个只读快照页，方便会后归档。"
        " 计划页应包含三名成员 `Alice`、`Bob`、`Cara` 的本周排班安排。"
        " 你可以自由决定实现步骤，但需要保证最终至少有一个主计划页和一个快照页，并自行确认工作簿结构。"
        ' 最终只输出一行 JSON，格式为 {"workbook":"team_plan_snapshot.xlsx","sheets":["Plan","Snapshot"],"members":["Alice","Bob","Cara"]}。'
    )


def _month_end_close_prompt(workdir: Path) -> str:
    workbook = workdir / "month_end_close.xlsx"
    return (
        f"请制作一个月结工作簿，保存到绝对路径 `{workbook}`。"
        " 业务目标是让财务同学能一眼看出收入、成本和最终利润，所以文件里至少要有收入、成本和汇总三个区域或工作表。"
        " 汇总区里需要保留 Excel 公式，而不是只填静态数字。"
        " 你可以自主规划 sheet 结构和操作顺序，但完成后必须自行读取并确认最终结果。"
        ' 最终只输出一行 JSON，格式为 {"workbook":"month_end_close.xlsx","sheets":["Revenue","Costs","Summary"],"formula_cells":["B2","B3"]}。'
    )


def _requests_register_prompt(workdir: Path) -> str:
    workbook = workdir / "requests_register.xlsx"
    return (
        f"请准备一个新的需求登记表，保存到绝对路径 `{workbook}`。"
        " 业务目标是让运营团队能直接把它当成 intake register 使用：表头清楚、数据区是原生 Excel table，里面至少要有三条请求记录。"
        " 你可以自由决定具体实现步骤，但结束前请自己检查表格结构是否已经可用。"
        ' 最终只输出一行 JSON，格式为 {"workbook":"requests_register.xlsx","sheet":"Requests","table_name":"RequestsTable","header":["Request","Owner","Priority","Status"]}。'
    )


def _board_packet_repair_prompt(workdir: Path) -> str:
    workbook = workdir / "board_packet_repair.xlsx"
    _seed_board_packet_draft(workbook)
    return (
        f"绝对路径 `{workbook}` 已经有一个董事会材料草稿，但结构还不适合直接分发。"
        " 请把草稿整理成可以给高层看的版本：把主页面改成更正式的概览页，再补一个细节页，并确保标题区更清晰。"
        " 你可以自己决定具体修复步骤，但结束前要检查工作簿结构。"
        ' 最终只输出一行 JSON，格式为 {"workbook":"board_packet_repair.xlsx","sheets":["Overview","Notes","Details"],"title_range":"A1:D1"}。'
    )


def _staging_cleanup_prompt(workdir: Path) -> str:
    workbook = workdir / "staging_cleanup.xlsx"
    _seed_staging_cleanup_draft(workbook)
    return (
        f"绝对路径 `{workbook}` 已有一个待清理的任务台账草稿。"
        " 业务目标是把里面的 staging 痕迹去掉，让最终表头从 A1 开始，且只保留 `Task`、`Owner`、`Status` 三列。"
        " 你可以自由决定如何修复，但不要重建一份全新的结构，应该基于现有草稿完成清理，并在结束前自己检查结果。"
        ' 最终只输出一行 JSON，格式为 {"workbook":"staging_cleanup.xlsx","sheet":"Backlog","header":["Task","Owner","Status"]}。'
    )


def _evaluate_executive_briefing(parsed: dict[str, Any] | None, workdir: Path) -> tuple[bool, bool, str]:
    workbook = workdir / "executive_briefing.xlsx"
    wb, wb_note = _load_workbook_safely(workbook)
    ws, ws_note = _sheet_or_note(wb, "Dashboard")
    if ws is None:
        answer_ok = (
            parsed is not None
            and parsed.get("workbook") == "executive_briefing.xlsx"
            and parsed.get("sheet") == "Dashboard"
            and sorted(str(item) for item in parsed.get("merged_ranges", [])) == ["A1:D1"]
            and [str(item) for item in parsed.get("metrics", [])] == ["Revenue", "Cost", "Profit"]
        )
        return answer_ok, False, ws_note or wb_note or "dashboard_unavailable"
    merged_ranges = sorted(str(item) for item in ws.merged_cells.ranges)
    metrics = [ws["A3"].value, ws["A4"].value, ws["A5"].value]
    workbook_ok = merged_ranges == ["A1:D1"] and metrics == ["Revenue", "Cost", "Profit"]
    answer_ok = (
        parsed is not None
        and parsed.get("workbook") == "executive_briefing.xlsx"
        and parsed.get("sheet") == "Dashboard"
        and sorted(str(item) for item in parsed.get("merged_ranges", [])) == ["A1:D1"]
        and [str(item) for item in parsed.get("metrics", [])] == ["Revenue", "Cost", "Profit"]
    )
    return answer_ok, workbook_ok, f"merged={merged_ranges} metrics={metrics}"


def _evaluate_team_plan_snapshot(parsed: dict[str, Any] | None, workdir: Path) -> tuple[bool, bool, str]:
    workbook = workdir / "team_plan_snapshot.xlsx"
    wb, wb_note = _load_workbook_safely(workbook)
    if wb is None:
        answer_ok = (
            parsed is not None
            and parsed.get("workbook") == "team_plan_snapshot.xlsx"
            and sorted(str(item) for item in parsed.get("sheets", [])) == ["Plan", "Snapshot"]
            and [str(item) for item in parsed.get("members", [])] == ["Alice", "Bob", "Cara"]
        )
        return answer_ok, False, wb_note or "workbook_unavailable"
    sheets = sorted(wb.sheetnames)
    ws, ws_note = _sheet_or_note(wb, "Plan")
    if ws is None:
        answer_ok = (
            parsed is not None
            and parsed.get("workbook") == "team_plan_snapshot.xlsx"
            and sorted(str(item) for item in parsed.get("sheets", [])) == ["Plan", "Snapshot"]
            and [str(item) for item in parsed.get("members", [])] == ["Alice", "Bob", "Cara"]
        )
        return answer_ok, False, ws_note or "plan_unavailable"
    members = [ws["A2"].value, ws["A3"].value, ws["A4"].value]
    workbook_ok = sheets == ["Plan", "Sheet1", "Snapshot"] and members == ["Alice", "Bob", "Cara"]
    answer_ok = (
        parsed is not None
        and parsed.get("workbook") == "team_plan_snapshot.xlsx"
        and sorted(str(item) for item in parsed.get("sheets", [])) == ["Plan", "Snapshot"]
        and [str(item) for item in parsed.get("members", [])] == ["Alice", "Bob", "Cara"]
    )
    return answer_ok, workbook_ok, f"sheets={sheets} members={members}"


def _evaluate_month_end_close(parsed: dict[str, Any] | None, workdir: Path) -> tuple[bool, bool, str]:
    workbook = workdir / "month_end_close.xlsx"
    wb, wb_note = _load_workbook_safely(workbook)
    if wb is None:
        answer_ok = (
            parsed is not None
            and parsed.get("workbook") == "month_end_close.xlsx"
            and sorted(str(item) for item in parsed.get("sheets", [])) == ["Costs", "Revenue", "Summary"]
            and [str(item) for item in parsed.get("formula_cells", [])] == ["B2", "B3"]
        )
        return answer_ok, False, wb_note or "workbook_unavailable"
    sheets = sorted(wb.sheetnames)
    ws, ws_note = _sheet_or_note(wb, "Summary")
    if ws is None:
        answer_ok = (
            parsed is not None
            and parsed.get("workbook") == "month_end_close.xlsx"
            and sorted(str(item) for item in parsed.get("sheets", [])) == ["Costs", "Revenue", "Summary"]
            and [str(item) for item in parsed.get("formula_cells", [])] == ["B2", "B3"]
        )
        return answer_ok, False, ws_note or "summary_unavailable"
    formulas = [ws["B2"].value, ws["B3"].value]
    workbook_ok = sheets == ["Costs", "Revenue", "Sheet1", "Summary"] and formulas == [
        "=Revenue!B2-Costs!B2",
        "=Revenue!B3-Costs!B3",
    ]
    answer_ok = (
        parsed is not None
        and parsed.get("workbook") == "month_end_close.xlsx"
        and sorted(str(item) for item in parsed.get("sheets", [])) == ["Costs", "Revenue", "Summary"]
        and [str(item) for item in parsed.get("formula_cells", [])] == ["B2", "B3"]
    )
    return answer_ok, workbook_ok, f"sheets={sheets} formulas={formulas}"


def _evaluate_requests_register(parsed: dict[str, Any] | None, workdir: Path) -> tuple[bool, bool, str]:
    workbook = workdir / "requests_register.xlsx"
    wb, wb_note = _load_workbook_safely(workbook)
    ws, ws_note = _sheet_or_note(wb, "Requests")
    if ws is None:
        answer_ok = (
            parsed is not None
            and parsed.get("workbook") == "requests_register.xlsx"
            and parsed.get("sheet") == "Requests"
            and parsed.get("table_name") == "RequestsTable"
            and [str(item) for item in parsed.get("header", [])] == ["Request", "Owner", "Priority", "Status"]
        )
        return answer_ok, False, ws_note or wb_note or "requests_unavailable"
    header = [ws["A1"].value, ws["B1"].value, ws["C1"].value, ws["D1"].value]
    table_names = sorted(ws.tables.keys())
    workbook_ok = header == ["Request", "Owner", "Priority", "Status"] and table_names == ["RequestsTable"]
    answer_ok = (
        parsed is not None
        and parsed.get("workbook") == "requests_register.xlsx"
        and parsed.get("sheet") == "Requests"
        and parsed.get("table_name") == "RequestsTable"
        and [str(item) for item in parsed.get("header", [])] == ["Request", "Owner", "Priority", "Status"]
    )
    return answer_ok, workbook_ok, f"header={header} tables={table_names}"


def _evaluate_board_packet_repair(parsed: dict[str, Any] | None, workdir: Path) -> tuple[bool, bool, str]:
    workbook = workdir / "board_packet_repair.xlsx"
    wb, wb_note = _load_workbook_safely(workbook)
    if wb is None:
        answer_ok = (
            parsed is not None
            and parsed.get("workbook") == "board_packet_repair.xlsx"
            and sorted(str(item) for item in parsed.get("sheets", [])) == ["Details", "Notes", "Overview"]
            and parsed.get("title_range") == "A1:D1"
        )
        return answer_ok, False, wb_note or "workbook_unavailable"
    sheets = sorted(wb.sheetnames)
    ws, ws_note = _sheet_or_note(wb, "Overview")
    if ws is None:
        answer_ok = (
            parsed is not None
            and parsed.get("workbook") == "board_packet_repair.xlsx"
            and sorted(str(item) for item in parsed.get("sheets", [])) == ["Details", "Notes", "Overview"]
            and parsed.get("title_range") == "A1:D1"
        )
        return answer_ok, False, ws_note or "overview_unavailable"
    merged_ranges = sorted(str(item) for item in ws.merged_cells.ranges)
    workbook_ok = sheets == ["Details", "Notes", "Overview"] and merged_ranges == ["A1:D1"]
    answer_ok = (
        parsed is not None
        and parsed.get("workbook") == "board_packet_repair.xlsx"
        and sorted(str(item) for item in parsed.get("sheets", [])) == ["Details", "Notes", "Overview"]
        and parsed.get("title_range") == "A1:D1"
    )
    return answer_ok, workbook_ok, f"sheets={sheets} merged={merged_ranges}"


def _evaluate_staging_cleanup(parsed: dict[str, Any] | None, workdir: Path) -> tuple[bool, bool, str]:
    workbook = workdir / "staging_cleanup.xlsx"
    wb, wb_note = _load_workbook_safely(workbook)
    ws, ws_note = _sheet_or_note(wb, "Backlog")
    if ws is None:
        answer_ok = (
            parsed is not None
            and parsed.get("workbook") == "staging_cleanup.xlsx"
            and parsed.get("sheet") == "Backlog"
            and [str(item) for item in parsed.get("header", [])] == ["Task", "Owner", "Status"]
        )
        return answer_ok, False, ws_note or wb_note or "backlog_unavailable"
    header = [ws["A1"].value, ws["B1"].value, ws["C1"].value]
    first_row = [ws["A2"].value, ws["B2"].value, ws["C2"].value]
    workbook_ok = header == ["Task", "Owner", "Status"] and first_row == ["Close books", "Mina", "Done"]
    answer_ok = (
        parsed is not None
        and parsed.get("workbook") == "staging_cleanup.xlsx"
        and parsed.get("sheet") == "Backlog"
        and [str(item) for item in parsed.get("header", [])] == ["Task", "Owner", "Status"]
    )
    return answer_ok, workbook_ok, f"header={header} row2={first_row}"


def _seed_board_packet_draft(workbook: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Draft"
    ws["A1"] = "Board Packet Draft"
    ws["A3"] = "Revenue"
    ws["B3"] = 1200
    ws["A4"] = "Cost"
    ws["B4"] = 760
    notes = wb.create_sheet("Notes")
    notes["A1"] = "Prepare speaker notes"
    workbook.parent.mkdir(parents=True, exist_ok=True)
    wb.save(workbook)


def _seed_staging_cleanup_draft(workbook: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Backlog"
    rows = [
        ["STAGING", "STAGING", "STAGING", "STAGING"],
        ["Temp", "Task", "Owner", "Status"],
        ["hold", "Close books", "Mina", "Done"],
        ["hold", "Prepare deck", "Jon", "In Progress"],
        ["hold", "Review KPIs", "Lia", "Pending"],
    ]
    for row in rows:
        ws.append(row)
    workbook.parent.mkdir(parents=True, exist_ok=True)
    wb.save(workbook)


__all__ = ["SuiteCase", "tier_d_autonomous_planning_cases", "tier_d_cases"]
