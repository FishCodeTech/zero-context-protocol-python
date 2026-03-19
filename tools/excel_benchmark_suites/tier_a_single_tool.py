from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from .common import ExcelBenchmarkCase


def tier_a_cases() -> list[ExcelBenchmarkCase]:
    return [
        ExcelBenchmarkCase("A", "tier_a_create_workbook", _create_workbook_prompt, {"create_workbook": 1}, _evaluate_create_workbook),
        ExcelBenchmarkCase("A", "tier_a_create_worksheet", _create_worksheet_prompt, {"create_worksheet": 1}, _evaluate_create_worksheet),
        ExcelBenchmarkCase("A", "tier_a_write_data", _write_data_prompt, {"write_data_to_excel": 1}, _evaluate_write_data),
        ExcelBenchmarkCase("A", "tier_a_read_data", _read_data_prompt, {"read_data_from_excel": 1}, _evaluate_read_data),
        ExcelBenchmarkCase("A", "tier_a_format_range", _format_range_prompt, {"format_range": 1}, _evaluate_format_range),
        ExcelBenchmarkCase("A", "tier_a_validate_formula_syntax", _validate_formula_prompt, {"validate_formula_syntax": 1}, _evaluate_validate_formula),
        ExcelBenchmarkCase("A", "tier_a_apply_formula", _apply_formula_prompt, {"apply_formula": 1}, _evaluate_apply_formula),
        ExcelBenchmarkCase("A", "tier_a_create_table", _create_table_prompt, {"create_table": 1}, _evaluate_create_table),
        ExcelBenchmarkCase("A", "tier_a_merge_cells", _merge_cells_prompt, {"merge_cells": 1}, _evaluate_merge_cells),
        ExcelBenchmarkCase("A", "tier_a_get_merged_cells", _get_merged_cells_prompt, {"get_merged_cells": 1}, _evaluate_get_merged_cells),
        ExcelBenchmarkCase("A", "tier_a_rename_worksheet", _rename_worksheet_prompt, {"rename_worksheet": 1}, _evaluate_rename_worksheet),
        ExcelBenchmarkCase("A", "tier_a_copy_worksheet", _copy_worksheet_prompt, {"copy_worksheet": 1}, _evaluate_copy_worksheet),
        ExcelBenchmarkCase("A", "tier_a_insert_rows", _insert_rows_prompt, {"insert_rows": 1}, _evaluate_insert_rows),
        ExcelBenchmarkCase("A", "tier_a_insert_columns", _insert_columns_prompt, {"insert_columns": 1}, _evaluate_insert_columns),
        ExcelBenchmarkCase("A", "tier_a_delete_sheet_rows", _delete_rows_prompt, {"delete_sheet_rows": 1}, _evaluate_delete_rows),
        ExcelBenchmarkCase("A", "tier_a_delete_sheet_columns", _delete_columns_prompt, {"delete_sheet_columns": 1}, _evaluate_delete_columns),
    ]


def _create_workbook_prompt(workdir: Path) -> str:
    workbook = workdir / "tier_a_create_workbook.xlsx"
    return (
        f"请在绝对路径 `{workbook}` 创建一个新的 Excel 工作簿。"
        ' 最终只输出一行 JSON，格式为 {"workbook":"tier_a_create_workbook.xlsx","active_sheet":"Sheet1"}。'
    )


def _create_worksheet_prompt(workdir: Path) -> str:
    workbook = workdir / "tier_a_create_worksheet.xlsx"
    _seed_blank_workbook(workbook)
    return (
        f"请在绝对路径 `{workbook}` 的工作簿中新增一个名为 `Review` 的工作表。"
        ' 最终只输出一行 JSON，格式为 {"workbook":"tier_a_create_worksheet.xlsx","sheet":"Review"}。'
    )


def _write_data_prompt(workdir: Path) -> str:
    workbook = workdir / "tier_a_write_data.xlsx"
    _seed_sheet(workbook, "Data")
    return (
        f"请在绝对路径 `{workbook}` 的 `Data` 工作表中，从 `A1` 开始写入两行数据："
        ' 第一行 `Name`,`Value`；第二行 `Latency`,`42`。'
        ' 最终只输出一行 JSON，格式为 {"workbook":"tier_a_write_data.xlsx","sheet":"Data","cell":"A2","value":"Latency"}。'
    )


def _read_data_prompt(workdir: Path) -> str:
    workbook = workdir / "tier_a_read_data.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Audit"
    for row in [["Metric", "Value"], ["Errors", 3], ["Warnings", 5]]:
        ws.append(row)
    workbook.parent.mkdir(parents=True, exist_ok=True)
    wb.save(workbook)
    return (
        f"请读取绝对路径 `{workbook}` 的 `Audit` 工作表中 `A1:B3` 的数据。"
        ' 最终只输出一行 JSON，格式为 {"workbook":"tier_a_read_data.xlsx","sheet":"Audit","rows":3,"header":["Metric","Value"]}。'
    )


def _format_range_prompt(workdir: Path) -> str:
    workbook = workdir / "tier_a_format_range.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws["A1"] = "Title"
    workbook.parent.mkdir(parents=True, exist_ok=True)
    wb.save(workbook)
    return (
        f"请把绝对路径 `{workbook}` 的 `Report!A1:C1` 设置为加粗并居中。"
        ' 最终只输出一行 JSON，格式为 {"workbook":"tier_a_format_range.xlsx","sheet":"Report","range":"A1:C1"}。'
    )


def _validate_formula_prompt(workdir: Path) -> str:
    workbook = workdir / "tier_a_validate_formula.xlsx"
    _seed_sheet(workbook, "Calc")
    return (
        f"请验证绝对路径 `{workbook}` 的 `Calc` 工作表中，把公式 `=B2+C2` 写入 `D2` 是否语法正确。"
        ' 最终只输出一行 JSON，格式为 {"workbook":"tier_a_validate_formula.xlsx","sheet":"Calc","cell":"D2","valid":true}。'
    )


def _apply_formula_prompt(workdir: Path) -> str:
    workbook = workdir / "tier_a_apply_formula.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Calc"
    for row in [["Item", "Q1", "Q2", "Total"], ["A", 10, 20, None]]:
        ws.append(row)
    workbook.parent.mkdir(parents=True, exist_ok=True)
    wb.save(workbook)
    return (
        f"请把绝对路径 `{workbook}` 的 `Calc!D2` 设置为公式 `=B2+C2`。"
        ' 最终只输出一行 JSON，格式为 {"workbook":"tier_a_apply_formula.xlsx","sheet":"Calc","cell":"D2"}。'
    )


def _create_table_prompt(workdir: Path) -> str:
    workbook = workdir / "tier_a_create_table.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    for row in [["Order", "Amount"], ["SO-1", 100], ["SO-2", 120]]:
        ws.append(row)
    workbook.parent.mkdir(parents=True, exist_ok=True)
    wb.save(workbook)
    return (
        f"请把绝对路径 `{workbook}` 的 `Orders!A1:B3` 创建为 Excel table，表名 `OrdersTable`。"
        ' 最终只输出一行 JSON，格式为 {"workbook":"tier_a_create_table.xlsx","sheet":"Orders","table_name":"OrdersTable"}。'
    )


def _merge_cells_prompt(workdir: Path) -> str:
    workbook = workdir / "tier_a_merge_cells.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    ws["A1"] = "Executive Summary"
    workbook.parent.mkdir(parents=True, exist_ok=True)
    wb.save(workbook)
    return (
        f"请把绝对路径 `{workbook}` 的 `Dashboard!A1:D1` 合并。"
        ' 最终只输出一行 JSON，格式为 {"workbook":"tier_a_merge_cells.xlsx","sheet":"Dashboard","merged":"A1:D1"}。'
    )


def _get_merged_cells_prompt(workdir: Path) -> str:
    workbook = workdir / "tier_a_get_merged_cells.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    ws["A1"] = "Executive Summary"
    ws.merge_cells("A1:D1")
    workbook.parent.mkdir(parents=True, exist_ok=True)
    wb.save(workbook)
    return (
        f"请查询绝对路径 `{workbook}` 的 `Dashboard` 工作表中的合并单元格。"
        ' 最终只输出一行 JSON，格式为 {"workbook":"tier_a_get_merged_cells.xlsx","sheet":"Dashboard","merged_ranges":["A1:D1"]}。'
    )


def _rename_worksheet_prompt(workdir: Path) -> str:
    workbook = workdir / "tier_a_rename_worksheet.xlsx"
    _seed_sheet(workbook, "Draft")
    return (
        f"请把绝对路径 `{workbook}` 里的工作表 `Draft` 重命名为 `Final`。"
        ' 最终只输出一行 JSON，格式为 {"workbook":"tier_a_rename_worksheet.xlsx","sheet":"Final"}。'
    )


def _copy_worksheet_prompt(workdir: Path) -> str:
    workbook = workdir / "tier_a_copy_worksheet.xlsx"
    _seed_sheet(workbook, "Template")
    return (
        f"请把绝对路径 `{workbook}` 中的工作表 `Template` 复制为 `Template Copy`。"
        ' 最终只输出一行 JSON，格式为 {"workbook":"tier_a_copy_worksheet.xlsx","sheet":"Template Copy"}。'
    )


def _insert_rows_prompt(workdir: Path) -> str:
    workbook = workdir / "tier_a_insert_rows.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Backlog"
    for row in [["Task"], ["A"], ["B"]]:
        ws.append(row)
    workbook.parent.mkdir(parents=True, exist_ok=True)
    wb.save(workbook)
    return (
        f"请在绝对路径 `{workbook}` 的 `Backlog` 工作表第 2 行插入一行。"
        ' 最终只输出一行 JSON，格式为 {"workbook":"tier_a_insert_rows.xlsx","sheet":"Backlog","inserted_before_row":2}。'
    )


def _insert_columns_prompt(workdir: Path) -> str:
    workbook = workdir / "tier_a_insert_columns.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Metrics"
    for row in [["Metric", "Actual"], ["Latency", 42]]:
        ws.append(row)
    workbook.parent.mkdir(parents=True, exist_ok=True)
    wb.save(workbook)
    return (
        f"请在绝对路径 `{workbook}` 的 `Metrics` 工作表第 2 列前插入一列。"
        ' 最终只输出一行 JSON，格式为 {"workbook":"tier_a_insert_columns.xlsx","sheet":"Metrics","inserted_before_column":2}。'
    )


def _delete_rows_prompt(workdir: Path) -> str:
    workbook = workdir / "tier_a_delete_rows.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Backlog"
    for row in [["Task"], ["A"], ["B"], ["C"]]:
        ws.append(row)
    workbook.parent.mkdir(parents=True, exist_ok=True)
    wb.save(workbook)
    return (
        f"请删除绝对路径 `{workbook}` 的 `Backlog` 工作表第 3 行。"
        ' 最终只输出一行 JSON，格式为 {"workbook":"tier_a_delete_rows.xlsx","sheet":"Backlog","remaining_rows":3}。'
    )


def _delete_columns_prompt(workdir: Path) -> str:
    workbook = workdir / "tier_a_delete_columns.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Metrics"
    for row in [["Metric", "Target", "Actual"], ["Latency", 40, 42]]:
        ws.append(row)
    workbook.parent.mkdir(parents=True, exist_ok=True)
    wb.save(workbook)
    return (
        f"请删除绝对路径 `{workbook}` 的 `Metrics` 工作表第 2 列。"
        ' 最终只输出一行 JSON，格式为 {"workbook":"tier_a_delete_columns.xlsx","sheet":"Metrics","remaining_header":["Metric","Actual"]}。'
    )


def _evaluate_create_workbook(parsed: dict[str, Any] | None, workdir: Path) -> tuple[bool, bool, str]:
    wb = load_workbook(workdir / "tier_a_create_workbook.xlsx")
    workbook_ok = wb.sheetnames == ["Sheet1"]
    answer_ok = parsed is not None and parsed.get("workbook") == "tier_a_create_workbook.xlsx" and parsed.get("active_sheet") == "Sheet1"
    return answer_ok, workbook_ok, f"sheets={wb.sheetnames}"


def _evaluate_create_worksheet(parsed: dict[str, Any] | None, workdir: Path) -> tuple[bool, bool, str]:
    wb = load_workbook(workdir / "tier_a_create_worksheet.xlsx")
    workbook_ok = sorted(wb.sheetnames) == ["Review", "Sheet1"]
    answer_ok = parsed is not None and parsed.get("workbook") == "tier_a_create_worksheet.xlsx" and parsed.get("sheet") == "Review"
    return answer_ok, workbook_ok, f"sheets={wb.sheetnames}"


def _evaluate_write_data(parsed: dict[str, Any] | None, workdir: Path) -> tuple[bool, bool, str]:
    ws = load_workbook(workdir / "tier_a_write_data.xlsx")["Data"]
    workbook_ok = ws["A2"].value == "Latency" and ws["B2"].value == 42
    answer_ok = parsed is not None and parsed.get("workbook") == "tier_a_write_data.xlsx" and parsed.get("sheet") == "Data" and parsed.get("cell") == "A2" and parsed.get("value") == "Latency"
    return answer_ok, workbook_ok, f"a2={ws['A2'].value} b2={ws['B2'].value}"


def _evaluate_read_data(parsed: dict[str, Any] | None, workdir: Path) -> tuple[bool, bool, str]:
    ws = load_workbook(workdir / "tier_a_read_data.xlsx")["Audit"]
    workbook_ok = ws.max_row == 3 and ws["A1"].value == "Metric"
    answer_ok = parsed is not None and parsed.get("workbook") == "tier_a_read_data.xlsx" and parsed.get("sheet") == "Audit" and int(parsed.get("rows", -1)) == 3 and parsed.get("header") == ["Metric", "Value"]
    return answer_ok, workbook_ok, f"rows={ws.max_row}"


def _evaluate_format_range(parsed: dict[str, Any] | None, workdir: Path) -> tuple[bool, bool, str]:
    ws = load_workbook(workdir / "tier_a_format_range.xlsx")["Report"]
    workbook_ok = bool(ws["A1"].font.bold) and ws["A1"].alignment.horizontal == "center"
    answer_ok = parsed is not None and parsed.get("workbook") == "tier_a_format_range.xlsx" and parsed.get("sheet") == "Report" and parsed.get("range") == "A1:C1"
    return answer_ok, workbook_ok, f"bold={ws['A1'].font.bold} align={ws['A1'].alignment.horizontal}"


def _evaluate_validate_formula(parsed: dict[str, Any] | None, workdir: Path) -> tuple[bool, bool, str]:
    workbook_ok = True
    answer_ok = parsed is not None and parsed.get("workbook") == "tier_a_validate_formula.xlsx" and parsed.get("sheet") == "Calc" and parsed.get("cell") == "D2" and bool(parsed.get("valid")) is True
    return answer_ok, workbook_ok, "validated"


def _evaluate_apply_formula(parsed: dict[str, Any] | None, workdir: Path) -> tuple[bool, bool, str]:
    ws = load_workbook(workdir / "tier_a_apply_formula.xlsx", data_only=False)["Calc"]
    workbook_ok = ws["D2"].value == "=B2+C2"
    answer_ok = parsed is not None and parsed.get("workbook") == "tier_a_apply_formula.xlsx" and parsed.get("sheet") == "Calc" and parsed.get("cell") == "D2"
    return answer_ok, workbook_ok, f"d2={ws['D2'].value}"


def _evaluate_create_table(parsed: dict[str, Any] | None, workdir: Path) -> tuple[bool, bool, str]:
    ws = load_workbook(workdir / "tier_a_create_table.xlsx")["Orders"]
    workbook_ok = "OrdersTable" in ws.tables
    answer_ok = parsed is not None and parsed.get("workbook") == "tier_a_create_table.xlsx" and parsed.get("sheet") == "Orders" and parsed.get("table_name") == "OrdersTable"
    return answer_ok, workbook_ok, f"tables={list(ws.tables.keys())}"


def _evaluate_merge_cells(parsed: dict[str, Any] | None, workdir: Path) -> tuple[bool, bool, str]:
    ws = load_workbook(workdir / "tier_a_merge_cells.xlsx")["Dashboard"]
    merged = sorted(str(item) for item in ws.merged_cells.ranges)
    workbook_ok = merged == ["A1:D1"]
    answer_ok = parsed is not None and parsed.get("workbook") == "tier_a_merge_cells.xlsx" and parsed.get("sheet") == "Dashboard" and parsed.get("merged") == "A1:D1"
    return answer_ok, workbook_ok, f"merged={merged}"


def _evaluate_get_merged_cells(parsed: dict[str, Any] | None, workdir: Path) -> tuple[bool, bool, str]:
    ws = load_workbook(workdir / "tier_a_get_merged_cells.xlsx")["Dashboard"]
    merged = sorted(str(item) for item in ws.merged_cells.ranges)
    workbook_ok = merged == ["A1:D1"]
    answer_ok = parsed is not None and parsed.get("workbook") == "tier_a_get_merged_cells.xlsx" and parsed.get("sheet") == "Dashboard" and parsed.get("merged_ranges") == ["A1:D1"]
    return answer_ok, workbook_ok, f"merged={merged}"


def _evaluate_rename_worksheet(parsed: dict[str, Any] | None, workdir: Path) -> tuple[bool, bool, str]:
    wb = load_workbook(workdir / "tier_a_rename_worksheet.xlsx")
    workbook_ok = sorted(wb.sheetnames) == ["Final", "Sheet1"]
    answer_ok = parsed is not None and parsed.get("workbook") == "tier_a_rename_worksheet.xlsx" and parsed.get("sheet") == "Final"
    return answer_ok, workbook_ok, f"sheets={wb.sheetnames}"


def _evaluate_copy_worksheet(parsed: dict[str, Any] | None, workdir: Path) -> tuple[bool, bool, str]:
    wb = load_workbook(workdir / "tier_a_copy_worksheet.xlsx")
    workbook_ok = sorted(wb.sheetnames) == ["Sheet1", "Template", "Template Copy"]
    answer_ok = parsed is not None and parsed.get("workbook") == "tier_a_copy_worksheet.xlsx" and parsed.get("sheet") == "Template Copy"
    return answer_ok, workbook_ok, f"sheets={wb.sheetnames}"


def _evaluate_insert_rows(parsed: dict[str, Any] | None, workdir: Path) -> tuple[bool, bool, str]:
    ws = load_workbook(workdir / "tier_a_insert_rows.xlsx")["Backlog"]
    workbook_ok = ws.max_row == 4 and ws["A3"].value == "A"
    answer_ok = parsed is not None and parsed.get("workbook") == "tier_a_insert_rows.xlsx" and parsed.get("sheet") == "Backlog" and int(parsed.get("inserted_before_row", -1)) == 2
    return answer_ok, workbook_ok, f"rows={ws.max_row}"


def _evaluate_insert_columns(parsed: dict[str, Any] | None, workdir: Path) -> tuple[bool, bool, str]:
    ws = load_workbook(workdir / "tier_a_insert_columns.xlsx")["Metrics"]
    workbook_ok = ws.max_column == 3 and ws["C1"].value == "Actual"
    answer_ok = parsed is not None and parsed.get("workbook") == "tier_a_insert_columns.xlsx" and parsed.get("sheet") == "Metrics" and int(parsed.get("inserted_before_column", -1)) == 2
    return answer_ok, workbook_ok, f"cols={ws.max_column}"


def _evaluate_delete_rows(parsed: dict[str, Any] | None, workdir: Path) -> tuple[bool, bool, str]:
    ws = load_workbook(workdir / "tier_a_delete_rows.xlsx")["Backlog"]
    workbook_ok = ws.max_row == 3 and [ws["A2"].value, ws["A3"].value] == ["A", "C"]
    answer_ok = parsed is not None and parsed.get("workbook") == "tier_a_delete_rows.xlsx" and parsed.get("sheet") == "Backlog" and int(parsed.get("remaining_rows", -1)) == 3
    return answer_ok, workbook_ok, f"rows={[ws['A1'].value, ws['A2'].value, ws['A3'].value]}"


def _evaluate_delete_columns(parsed: dict[str, Any] | None, workdir: Path) -> tuple[bool, bool, str]:
    ws = load_workbook(workdir / "tier_a_delete_columns.xlsx")["Metrics"]
    header = [ws["A1"].value, ws["B1"].value]
    workbook_ok = ws.max_column == 2 and header == ["Metric", "Actual"]
    answer_ok = parsed is not None and parsed.get("workbook") == "tier_a_delete_columns.xlsx" and parsed.get("sheet") == "Metrics" and parsed.get("remaining_header") == ["Metric", "Actual"]
    return answer_ok, workbook_ok, f"header={header}"


def _seed_blank_workbook(workbook: Path) -> None:
    wb = Workbook()
    wb.active.title = "Sheet1"
    workbook.parent.mkdir(parents=True, exist_ok=True)
    wb.save(workbook)


def _seed_sheet(workbook: Path, sheet_name: str) -> None:
    wb = Workbook()
    wb.active.title = "Sheet1"
    wb.create_sheet(sheet_name)
    workbook.parent.mkdir(parents=True, exist_ok=True)
    wb.save(workbook)
