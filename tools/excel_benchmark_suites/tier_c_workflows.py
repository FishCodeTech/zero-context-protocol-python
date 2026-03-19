from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook

from .common import ExcelBenchmarkCase


@dataclass(frozen=True)
class TierWorkflowCase:
    case_id: str
    prompt_factory: Callable[[Path], str]
    expected_tool_calls: dict[str, int]
    evaluator: Callable[[dict[str, Any] | None, Path], tuple[bool, bool, str]]


def workflow_cases() -> list[TierWorkflowCase]:
    return [
        TierWorkflowCase(
            case_id="sales_ops_weekly_pack",
            prompt_factory=_sales_ops_weekly_pack_prompt,
            expected_tool_calls={
                "create_workbook": 1,
                "create_worksheet": 2,
                "write_data_to_excel": 2,
                "create_table": 1,
                "format_range": 1,
                "apply_formula": 2,
                "read_data_from_excel": 1,
            },
            evaluator=_evaluate_sales_ops_weekly_pack,
        ),
        TierWorkflowCase(
            case_id="finance_close_summary",
            prompt_factory=_finance_close_summary_prompt,
            expected_tool_calls={
                "create_workbook": 1,
                "create_worksheet": 1,
                "write_data_to_excel": 1,
                "validate_formula_syntax": 1,
                "apply_formula": 3,
                "format_range": 1,
                "create_table": 1,
                "read_data_from_excel": 1,
            },
            evaluator=_evaluate_finance_close_summary,
        ),
        TierWorkflowCase(
            case_id="kpi_dashboard_sheet",
            prompt_factory=_kpi_dashboard_sheet_prompt,
            expected_tool_calls={
                "create_workbook": 1,
                "create_worksheet": 1,
                "write_data_to_excel": 2,
                "merge_cells": 1,
                "format_range": 2,
                "get_merged_cells": 1,
                "read_data_from_excel": 1,
            },
            evaluator=_evaluate_kpi_dashboard_sheet,
        ),
        TierWorkflowCase(
            case_id="multi_sheet_monthly_rollup",
            prompt_factory=_multi_sheet_monthly_rollup_prompt,
            expected_tool_calls={
                "create_workbook": 1,
                "create_worksheet": 3,
                "write_data_to_excel": 3,
                "apply_formula": 2,
                "get_workbook_metadata": 1,
                "read_data_from_excel": 1,
            },
            evaluator=_evaluate_multi_sheet_monthly_rollup,
        ),
        TierWorkflowCase(
            case_id="template_copy_and_fill",
            prompt_factory=_template_copy_and_fill_prompt,
            expected_tool_calls={
                "create_workbook": 1,
                "create_worksheet": 1,
                "write_data_to_excel": 3,
                "copy_worksheet": 2,
                "rename_worksheet": 1,
                "read_data_from_excel": 1,
            },
            evaluator=_evaluate_template_copy_and_fill,
        ),
        TierWorkflowCase(
            case_id="headcount_plan_restructure",
            prompt_factory=_headcount_plan_restructure_prompt,
            expected_tool_calls={
                "create_workbook": 1,
                "create_worksheet": 1,
                "write_data_to_excel": 1,
                "insert_rows": 1,
                "insert_columns": 1,
                "delete_sheet_rows": 1,
                "delete_sheet_columns": 1,
                "read_data_from_excel": 1,
            },
            evaluator=_evaluate_headcount_plan_restructure,
        ),
        TierWorkflowCase(
            case_id="revenue_model_revision",
            prompt_factory=_revenue_model_revision_prompt,
            expected_tool_calls={
                "create_workbook": 1,
                "create_worksheet": 1,
                "write_data_to_excel": 1,
                "apply_formula": 2,
                "format_range": 1,
                "merge_cells": 1,
                "read_data_from_excel": 1,
            },
            evaluator=_evaluate_revenue_model_revision,
        ),
    ]


def tier_c_cases() -> list[ExcelBenchmarkCase]:
    semantic_tools = {
        "sales_ops_weekly_pack": {"zcp_build_sales_ops_weekly_pack": 1},
        "finance_close_summary": {"zcp_finalize_finance_close_summary": 1},
        "kpi_dashboard_sheet": {"zcp_build_kpi_dashboard_sheet": 1},
        "multi_sheet_monthly_rollup": {"zcp_build_multi_sheet_monthly_rollup": 1},
        "template_copy_and_fill": {"zcp_apply_template_copy_and_fill": 1},
        "headcount_plan_restructure": {"zcp_restructure_headcount_plan": 1},
        "revenue_model_revision": {"zcp_revise_revenue_model": 1},
    }
    return [
        ExcelBenchmarkCase(
            tier="C",
            case_id=case.case_id,
            prompt_factory=case.prompt_factory,
            required_tool_calls=case.expected_tool_calls,
            evaluator=case.evaluator,
            native_zcp_required_tool_calls=semantic_tools[case.case_id],
        )
        for case in workflow_cases()
    ]


def _sales_ops_weekly_pack_prompt(workdir: Path) -> str:
    workbook = workdir / "sales_ops_weekly_pack.xlsx"
    return (
        f"请在绝对路径 `{workbook}` 创建一个 Excel 工作簿，包含 `Pipeline` 和 `Summary` 两个工作表。"
        " 在 `Pipeline!A1:E5` 写入表头和 4 行商机数据，列为 Deal、Owner、Stage、Amount、CloseWeek。"
        " 将 `Pipeline!A1:E5` 创建为表格 `PipelineTable`，并把表头设为加粗。"
        " 在 `Summary!A1:B3` 写入 `Metric/Value`、`Qualified Deals/`、`Pipeline Amount/` 结构，"
        " 再用公式把 `B2` 设为 `=COUNTA(Pipeline!A2:A5)`，把 `B3` 设为 `=SUM(Pipeline!D2:D5)`。"
        " 最后读取 `Summary!A1:B3` 进行确认。"
        ' 最终只输出一行 JSON，格式为 {"workbook":"sales_ops_weekly_pack.xlsx","sheets":["Pipeline","Sheet1","Summary"],"summary_formula_cells":["B2","B3"],"table_name":"PipelineTable"}。'
    )


def _finance_close_summary_prompt(workdir: Path) -> str:
    workbook = workdir / "finance_close_summary.xlsx"
    return (
        f"请在绝对路径 `{workbook}` 创建工作簿，并新增工作表 `ClosePack`。"
        " 在 `ClosePack!A1:D5` 写入财务关账数据，列为 Category、Jan、Feb、Q1Total，"
        " 行为 Revenue/420/460、COGS/180/190、Opex/120/130、GrossProfit/null/null。"
        " 先验证公式语法，再把 `D2` 设为 `=B2+C2`、`D3` 设为 `=B3+C3`、`D4` 设为 `=B4+C4`。"
        " 将 `A1:D4` 创建为表格 `CloseTable`，并把 `A1:D1` 设为加粗和浅灰背景。"
        " 最后读取 `A1:D4`。"
        ' 最终只输出一行 JSON，格式为 {"workbook":"finance_close_summary.xlsx","sheet":"ClosePack","table_name":"CloseTable","formula_cells":["D2","D3","D4"]}。'
    )


def _kpi_dashboard_sheet_prompt(workdir: Path) -> str:
    workbook = workdir / "kpi_dashboard_sheet.xlsx"
    return (
        f"请在绝对路径 `{workbook}` 创建工作簿，并新增工作表 `Dashboard`。"
        " 在 `A1` 写入标题 `Executive KPI Dashboard`，然后把 `A1:D1` 合并。"
        " 在 `A3:B6` 写入 4 行 KPI：Revenue/1280、GrossMargin/0.42、NPS/61、Churn/0.03。"
        " 把标题区域 `A1:D1` 设为加粗和居中，把 KPI 表头 `A3:B3` 设为加粗。"
        " 查询合并单元格，并读取 `A1:B6`。"
        ' 最终只输出一行 JSON，格式为 {"workbook":"kpi_dashboard_sheet.xlsx","sheet":"Dashboard","merged_ranges":["A1:D1"],"metrics":["Revenue","GrossMargin","NPS","Churn"]}。'
    )


def _multi_sheet_monthly_rollup_prompt(workdir: Path) -> str:
    workbook = workdir / "multi_sheet_monthly_rollup.xlsx"
    return (
        f"请在绝对路径 `{workbook}` 创建工作簿，新增 `Jan`、`Feb`、`Summary` 三个工作表。"
        " 在 `Jan!A1:B3` 写入 Product/Revenue 和两行 Alpha/240、Beta/190。"
        " 在 `Feb!A1:B3` 写入 Product/Revenue 和两行 Alpha/260、Beta/210。"
        " 在 `Summary!A1:B3` 写入 Product/TotalRevenue 和两行 Alpha/null、Beta/null。"
        " 用公式把 `Summary!B2` 设为 `=Jan!B2+Feb!B2`，把 `Summary!B3` 设为 `=Jan!B3+Feb!B3`。"
        " 获取工作簿元数据，并读取 `Summary!A1:B3`。"
        ' 最终只输出一行 JSON，格式为 {"workbook":"multi_sheet_monthly_rollup.xlsx","sheets":["Feb","Jan","Sheet1","Summary"],"summary_formula_cells":["B2","B3"]}。'
    )


def _template_copy_and_fill_prompt(workdir: Path) -> str:
    workbook = workdir / "template_copy_and_fill.xlsx"
    return (
        f"请在绝对路径 `{workbook}` 创建工作簿，并新增工作表 `Template`。"
        " 在 `Template!A1:C3` 写入表头和占位内容，表头为 Client、Owner、Quota。"
        " 复制 `Template` 两次，得到 `NorthRegion` 和 `SouthRegion`。"
        " 再把原来的 `Template` 重命名为 `MasterTemplate`。"
        " 在 `NorthRegion!A2:C2` 写入 North/Casey/180，在 `SouthRegion!A2:C2` 写入 South/Jordan/210。"
        " 最后读取 `NorthRegion!A1:C2`。"
        ' 最终只输出一行 JSON，格式为 {"workbook":"template_copy_and_fill.xlsx","sheets":["MasterTemplate","NorthRegion","Sheet1","SouthRegion"],"filled_sheet":"NorthRegion","quota":180}。'
    )


def _headcount_plan_restructure_prompt(workdir: Path) -> str:
    workbook = workdir / "headcount_plan_restructure.xlsx"
    return (
        f"请在绝对路径 `{workbook}` 创建工作簿，并新增工作表 `Plan`。"
        " 在 `Plan!A1:C5` 写入 Dept、CurrentHC、TargetHC 和 4 行数据：Sales/12/15、CS/9/10、Ops/7/8、Temp/2/0。"
        " 在第 2 行下方插入一行，用于新部门 `Marketing/5/7`；"
        " 再在 B 列前插入一列，列名写为 `Region`，并把 Sales 和 Marketing 的 Region 分别写为 East 和 West。"
        " 删除最后的 `Temp` 行，并删除多余的最后一列。"
        " 最后读取 `A1:D6`。"
        ' 最终只输出一行 JSON，格式为 {"workbook":"headcount_plan_restructure.xlsx","sheet":"Plan","departments":["Sales","Marketing","CS","Ops"],"region_column":"B"}。'
    )


def _revenue_model_revision_prompt(workdir: Path) -> str:
    workbook = workdir / "revenue_model_revision.xlsx"
    return (
        f"请在绝对路径 `{workbook}` 创建工作簿，并新增工作表 `Model`。"
        " 在 `A1:D4` 写入表头 Segment、ARR、Expansion、Total 和三行数据 SMB/320/40/null、MidMarket/470/55/null、Enterprise/610/90/null。"
        " 将 `A1:D1` 设为加粗，并把 `A1:D1` 上方增加一个标题区域：在 `A1` 写入 `Revenue Model Review` 后把 `A1:D1` 合并。"
        " 然后把原数据表从第 2 行开始保留，并用公式把 `D3` 设为 `=B3+C3`，`D4` 设为 `=B4+C4`。"
        " 最后读取 `A1:D5`。"
        ' 最终只输出一行 JSON，格式为 {"workbook":"revenue_model_revision.xlsx","sheet":"Model","merged_ranges":["A1:D1"],"formula_cells":["D3","D4"]}。'
    )


def _evaluate_sales_ops_weekly_pack(parsed: dict[str, Any] | None, workdir: Path) -> tuple[bool, bool, str]:
    wb = load_workbook(workdir / "sales_ops_weekly_pack.xlsx", data_only=False)
    pipeline = wb["Pipeline"]
    summary = wb["Summary"]
    workbook_ok = (
        sorted(wb.sheetnames) == ["Pipeline", "Sheet1", "Summary"]
        and "PipelineTable" in pipeline.tables
        and summary["B2"].value == "=COUNTA(Pipeline!A2:A5)"
        and summary["B3"].value == "=SUM(Pipeline!D2:D5)"
    )
    answer_ok = (
        parsed is not None
        and parsed.get("workbook") == "sales_ops_weekly_pack.xlsx"
        and sorted(parsed.get("sheets", [])) == ["Pipeline", "Sheet1", "Summary"]
        and parsed.get("summary_formula_cells") == ["B2", "B3"]
        and parsed.get("table_name") == "PipelineTable"
    )
    return answer_ok, workbook_ok, f"sheets={wb.sheetnames} tables={list(pipeline.tables.keys())}"


def _evaluate_finance_close_summary(parsed: dict[str, Any] | None, workdir: Path) -> tuple[bool, bool, str]:
    wb = load_workbook(workdir / "finance_close_summary.xlsx", data_only=False)
    sheet = wb["ClosePack"]
    formulas = [sheet["D2"].value, sheet["D3"].value, sheet["D4"].value]
    workbook_ok = "CloseTable" in sheet.tables and formulas == ["=B2+C2", "=B3+C3", "=B4+C4"]
    answer_ok = (
        parsed is not None
        and parsed.get("workbook") == "finance_close_summary.xlsx"
        and parsed.get("sheet") == "ClosePack"
        and parsed.get("table_name") == "CloseTable"
        and parsed.get("formula_cells") == ["D2", "D3", "D4"]
    )
    return answer_ok, workbook_ok, f"formulas={formulas} tables={list(sheet.tables.keys())}"


def _evaluate_kpi_dashboard_sheet(parsed: dict[str, Any] | None, workdir: Path) -> tuple[bool, bool, str]:
    wb = load_workbook(workdir / "kpi_dashboard_sheet.xlsx")
    sheet = wb["Dashboard"]
    merged = sorted(str(item) for item in sheet.merged_cells.ranges)
    metrics = [sheet["A3"].value, sheet["A4"].value, sheet["A5"].value, sheet["A6"].value]
    workbook_ok = merged == ["A1:D1"] and metrics == ["Revenue", "GrossMargin", "NPS", "Churn"]
    answer_ok = (
        parsed is not None
        and parsed.get("workbook") == "kpi_dashboard_sheet.xlsx"
        and parsed.get("sheet") == "Dashboard"
        and parsed.get("merged_ranges") == ["A1:D1"]
        and parsed.get("metrics") == ["Revenue", "GrossMargin", "NPS", "Churn"]
    )
    return answer_ok, workbook_ok, f"merged={merged} metrics={metrics}"


def _evaluate_multi_sheet_monthly_rollup(parsed: dict[str, Any] | None, workdir: Path) -> tuple[bool, bool, str]:
    wb = load_workbook(workdir / "multi_sheet_monthly_rollup.xlsx", data_only=False)
    summary = wb["Summary"]
    formulas = [summary["B2"].value, summary["B3"].value]
    workbook_ok = sorted(wb.sheetnames) == ["Feb", "Jan", "Sheet1", "Summary"] and formulas == [
        "=Jan!B2+Feb!B2",
        "=Jan!B3+Feb!B3",
    ]
    answer_ok = (
        parsed is not None
        and parsed.get("workbook") == "multi_sheet_monthly_rollup.xlsx"
        and sorted(parsed.get("sheets", [])) == ["Feb", "Jan", "Sheet1", "Summary"]
        and parsed.get("summary_formula_cells") == ["B2", "B3"]
    )
    return answer_ok, workbook_ok, f"sheets={wb.sheetnames} formulas={formulas}"


def _evaluate_template_copy_and_fill(parsed: dict[str, Any] | None, workdir: Path) -> tuple[bool, bool, str]:
    wb = load_workbook(workdir / "template_copy_and_fill.xlsx")
    north = wb["NorthRegion"]
    workbook_ok = sorted(wb.sheetnames) == ["MasterTemplate", "NorthRegion", "Sheet1", "SouthRegion"] and north["A2"].value == "North" and north["C2"].value == 180
    answer_ok = (
        parsed is not None
        and parsed.get("workbook") == "template_copy_and_fill.xlsx"
        and sorted(parsed.get("sheets", [])) == ["MasterTemplate", "NorthRegion", "Sheet1", "SouthRegion"]
        and parsed.get("filled_sheet") == "NorthRegion"
        and int(parsed.get("quota", -1)) == 180
    )
    return answer_ok, workbook_ok, f"sheets={wb.sheetnames} north_row={[north['A2'].value, north['B2'].value, north['C2'].value]}"


def _evaluate_headcount_plan_restructure(parsed: dict[str, Any] | None, workdir: Path) -> tuple[bool, bool, str]:
    wb = load_workbook(workdir / "headcount_plan_restructure.xlsx")
    sheet = wb["Plan"]
    departments = [sheet["A2"].value, sheet["A3"].value, sheet["A4"].value, sheet["A5"].value]
    workbook_ok = departments == ["Sales", "Marketing", "CS", "Ops"] and sheet["B1"].value == "Region"
    answer_ok = (
        parsed is not None
        and parsed.get("workbook") == "headcount_plan_restructure.xlsx"
        and parsed.get("sheet") == "Plan"
        and parsed.get("departments") == ["Sales", "Marketing", "CS", "Ops"]
        and parsed.get("region_column") == "B"
    )
    return answer_ok, workbook_ok, f"departments={departments} headers={[sheet['A1'].value, sheet['B1'].value, sheet['C1'].value, sheet['D1'].value]}"


def _evaluate_revenue_model_revision(parsed: dict[str, Any] | None, workdir: Path) -> tuple[bool, bool, str]:
    wb = load_workbook(workdir / "revenue_model_revision.xlsx", data_only=False)
    sheet = wb["Model"]
    merged = sorted(str(item) for item in sheet.merged_cells.ranges)
    formulas = [sheet["D3"].value, sheet["D4"].value]
    workbook_ok = merged == ["A1:D1"] and formulas == ["=B3+C3", "=B4+C4"]
    answer_ok = (
        parsed is not None
        and parsed.get("workbook") == "revenue_model_revision.xlsx"
        and parsed.get("sheet") == "Model"
        and parsed.get("merged_ranges") == ["A1:D1"]
        and parsed.get("formula_cells") == ["D3", "D4"]
    )
    return answer_ok, workbook_ok, f"merged={merged} formulas={formulas}"
