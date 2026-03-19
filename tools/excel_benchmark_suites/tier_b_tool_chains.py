from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable

from openpyxl import Workbook, load_workbook

from .common import ExcelBenchmarkCase


@dataclass(frozen=True)
class TierBCaseSpec:
    case_id: str
    prompt_factory: Callable[[Path], str]
    expected_tool_calls: dict[str, int]
    evaluator: Callable[[dict[str, Any] | None, Path], tuple[bool, bool, str]]
    tier: str = "B"


def build_tier_b_cases(case_cls: type | None = None) -> list[Any]:
    case_type = case_cls or TierBCaseSpec
    return [
        case_type(
            case_id="tier_b_workbook_bootstrap_chain",
            prompt_factory=_bootstrap_chain_prompt,
            expected_tool_calls={
                "create_workbook": 1,
                "create_worksheet": 2,
                "get_workbook_metadata": 1,
            },
            evaluator=_evaluate_bootstrap_chain,
        ),
        case_type(
            case_id="tier_b_formula_flow_chain",
            prompt_factory=_formula_flow_prompt,
            expected_tool_calls={
                "validate_formula_syntax": 1,
                "apply_formula": 2,
                "read_data_from_excel": 1,
            },
            evaluator=_evaluate_formula_flow,
        ),
        case_type(
            case_id="tier_b_layout_flow_chain",
            prompt_factory=_layout_flow_prompt,
            expected_tool_calls={
                "merge_cells": 1,
                "format_range": 1,
                "get_merged_cells": 1,
            },
            evaluator=_evaluate_layout_flow,
        ),
        case_type(
            case_id="tier_b_sheet_maintenance_chain",
            prompt_factory=_sheet_maintenance_prompt,
            expected_tool_calls={
                "rename_worksheet": 1,
                "copy_worksheet": 1,
                "get_workbook_metadata": 1,
            },
            evaluator=_evaluate_sheet_maintenance,
        ),
        case_type(
            case_id="tier_b_table_setup_chain",
            prompt_factory=_table_setup_prompt,
            expected_tool_calls={
                "create_table": 1,
                "format_range": 1,
                "read_data_from_excel": 1,
            },
            evaluator=_evaluate_table_setup,
        ),
        case_type(
            case_id="tier_b_readback_verification_chain",
            prompt_factory=_readback_verification_prompt,
            expected_tool_calls={
                "validate_excel_range": 1,
                "read_data_from_excel": 1,
                "get_workbook_metadata": 1,
            },
            evaluator=_evaluate_readback_verification,
        ),
        case_type(
            case_id="tier_b_row_maintenance_chain",
            prompt_factory=_row_maintenance_prompt,
            expected_tool_calls={
                "insert_rows": 1,
                "write_data_to_excel": 1,
                "delete_sheet_rows": 1,
                "read_data_from_excel": 1,
            },
            evaluator=_evaluate_row_maintenance,
        ),
        case_type(
            case_id="tier_b_column_maintenance_chain",
            prompt_factory=_column_maintenance_prompt,
            expected_tool_calls={
                "insert_columns": 1,
                "write_data_to_excel": 1,
                "delete_sheet_columns": 1,
                "read_data_from_excel": 1,
            },
            evaluator=_evaluate_column_maintenance,
        ),
    ]


def tier_b_cases() -> list[ExcelBenchmarkCase]:
    semantic_tools = {
        "tier_b_workbook_bootstrap_chain": {"zcp_chain_workbook_bootstrap": 1},
        "tier_b_formula_flow_chain": {"zcp_chain_formula_flow": 1},
        "tier_b_layout_flow_chain": {"zcp_chain_layout_flow": 1},
        "tier_b_sheet_maintenance_chain": {"zcp_chain_sheet_maintenance": 1},
        "tier_b_table_setup_chain": {"zcp_chain_table_setup": 1},
        "tier_b_readback_verification_chain": {"zcp_chain_readback_verification": 1},
        "tier_b_row_maintenance_chain": {"zcp_chain_row_maintenance": 1},
        "tier_b_column_maintenance_chain": {"zcp_chain_column_maintenance": 1},
    }
    return [
        ExcelBenchmarkCase(
            tier="B",
            case_id=case.case_id,
            prompt_factory=case.prompt_factory,
            required_tool_calls=case.expected_tool_calls,
            evaluator=case.evaluator,
            native_zcp_required_tool_calls=semantic_tools[case.case_id],
        )
        for case in build_tier_b_cases()
    ]


def _bootstrap_chain_prompt(workdir: Path) -> str:
    workbook = workdir / "tier_b_bootstrap.xlsx"
    return (
        f"请在绝对路径 `{workbook}` 创建工作簿，并新增 `RawData` 与 `Summary` 两个工作表。"
        " 然后读取 workbook metadata 确认最终工作表集合。"
        ' 最终只输出一行 JSON，格式为 {"workbook":"tier_b_bootstrap.xlsx","sheets":["RawData","Sheet1","Summary"],"sheet_count":3}。'
    )


def _formula_flow_prompt(workdir: Path) -> str:
    workbook = _seed_formula_flow_workbook(workdir)
    return (
        f"请处理绝对路径 `{workbook}` 的 `Sales` 工作表。"
        " 先验证 `=B2+C2` 的公式语法，再把 `=B2+C2` 和 `=B3+C3` 分别写入 `D2` 与 `D3`，"
        " 最后读取 `A1:D3` 确认结果。"
        ' 最终只输出一行 JSON，格式为 {"workbook":"tier_b_formula_flow.xlsx","sheet":"Sales","formula_cells":["D2","D3"],"header":["Item","Q1","Q2","Total"]}。'
    )


def _layout_flow_prompt(workdir: Path) -> str:
    workbook = _seed_layout_flow_workbook(workdir)
    return (
        f"请处理绝对路径 `{workbook}` 的 `Dashboard` 工作表。"
        " 将 `A1:D1` 合并，把标题行设置为加粗并居中，然后查询合并单元格。"
        ' 最终只输出一行 JSON，格式为 {"workbook":"tier_b_layout_flow.xlsx","sheet":"Dashboard","merged_ranges":["A1:D1"],"title":"Operations Dashboard"}。'
    )


def _sheet_maintenance_prompt(workdir: Path) -> str:
    workbook = _seed_sheet_maintenance_workbook(workdir)
    return (
        f"请处理绝对路径 `{workbook}` 的 workbook。"
        " 将工作表 `Draft` 重命名为 `Template`，复制 `Template` 为 `Template Copy`，然后读取 workbook metadata。"
        ' 最终只输出一行 JSON，格式为 {"workbook":"tier_b_sheet_maintenance.xlsx","sheets":["Sheet1","Template","Template Copy"]}。'
    )


def _table_setup_prompt(workdir: Path) -> str:
    workbook = _seed_table_setup_workbook(workdir)
    return (
        f"请处理绝对路径 `{workbook}` 的 `Orders` 工作表。"
        " 将 `A1:C4` 创建为 Excel table，表名为 `OrdersTable`，然后把表头 `A1:C1` 设为加粗，最后读取 `A1:C4`。"
        ' 最终只输出一行 JSON，格式为 {"workbook":"tier_b_table_setup.xlsx","sheet":"Orders","table_name":"OrdersTable","header":["Order","Amount","Status"]}。'
    )


def _readback_verification_prompt(workdir: Path) -> str:
    workbook = _seed_readback_verification_workbook(workdir)
    return (
        f"请处理绝对路径 `{workbook}` 的 `Audit` 工作表。"
        " 先验证 `A1:C4` 是有效范围，再读取 `A1:C4`，然后读取 workbook metadata 确认使用范围。"
        ' 最终只输出一行 JSON，格式为 {"workbook":"tier_b_readback.xlsx","sheet":"Audit","range":"A1:C4","used_range":"A1:C4"}。'
    )


def _row_maintenance_prompt(workdir: Path) -> str:
    workbook = _seed_row_maintenance_workbook(workdir)
    return (
        f"请处理绝对路径 `{workbook}` 的 `Backlog` 工作表。"
        " 在第 3 行插入一行，在 `A3:B3` 写入 `Hotfix` 和 `In Progress`，然后删除原来的第 5 行，最后读取 `A1:B4`。"
        ' 最终只输出一行 JSON，格式为 {"workbook":"tier_b_row_ops.xlsx","sheet":"Backlog","rows":["Task A","Hotfix","Task B"]}。'
    )


def _column_maintenance_prompt(workdir: Path) -> str:
    workbook = _seed_column_maintenance_workbook(workdir)
    return (
        f"请处理绝对路径 `{workbook}` 的 `Metrics` 工作表。"
        " 在第 2 列插入一列，在 `B1:B3` 写入 `Target`、`95`、`88`，然后删除第 4 列，最后读取 `A1:C3`。"
        ' 最终只输出一行 JSON，格式为 {"workbook":"tier_b_column_ops.xlsx","sheet":"Metrics","header":["Metric","Target","Actual"]}。'
    )


def _evaluate_bootstrap_chain(parsed: dict[str, Any] | None, workdir: Path) -> tuple[bool, bool, str]:
    workbook = workdir / "tier_b_bootstrap.xlsx"
    wb = load_workbook(workbook)
    sheets = sorted(wb.sheetnames)
    workbook_ok = sheets == ["RawData", "Sheet1", "Summary"]
    answer_ok = (
        parsed is not None
        and str(parsed.get("workbook")) == "tier_b_bootstrap.xlsx"
        and int(parsed.get("sheet_count", -1)) == 3
        and sorted(str(item) for item in parsed.get("sheets", [])) == sheets
    )
    return answer_ok, workbook_ok, f"sheets={sheets}"


def _evaluate_formula_flow(parsed: dict[str, Any] | None, workdir: Path) -> tuple[bool, bool, str]:
    workbook = workdir / "tier_b_formula_flow.xlsx"
    ws = load_workbook(workbook, data_only=False)["Sales"]
    formulas = [ws["D2"].value, ws["D3"].value]
    header = [ws["A1"].value, ws["B1"].value, ws["C1"].value, ws["D1"].value]
    workbook_ok = formulas == ["=B2+C2", "=B3+C3"] and header == ["Item", "Q1", "Q2", "Total"]
    answer_ok = (
        parsed is not None
        and str(parsed.get("workbook")) == "tier_b_formula_flow.xlsx"
        and str(parsed.get("sheet")) == "Sales"
        and [str(item) for item in parsed.get("formula_cells", [])] == ["D2", "D3"]
        and [str(item) for item in parsed.get("header", [])] == ["Item", "Q1", "Q2", "Total"]
    )
    return answer_ok, workbook_ok, f"formulas={formulas}"


def _evaluate_layout_flow(parsed: dict[str, Any] | None, workdir: Path) -> tuple[bool, bool, str]:
    workbook = workdir / "tier_b_layout_flow.xlsx"
    ws = load_workbook(workbook)["Dashboard"]
    merged = sorted(str(item) for item in ws.merged_cells.ranges)
    title = ws["A1"].value
    workbook_ok = merged == ["A1:D1"] and title == "Operations Dashboard"
    answer_ok = (
        parsed is not None
        and str(parsed.get("workbook")) == "tier_b_layout_flow.xlsx"
        and str(parsed.get("sheet")) == "Dashboard"
        and sorted(str(item) for item in parsed.get("merged_ranges", [])) == ["A1:D1"]
        and str(parsed.get("title")) == "Operations Dashboard"
    )
    return answer_ok, workbook_ok, f"merged={merged}"


def _evaluate_sheet_maintenance(parsed: dict[str, Any] | None, workdir: Path) -> tuple[bool, bool, str]:
    workbook = workdir / "tier_b_sheet_maintenance.xlsx"
    sheets = sorted(load_workbook(workbook).sheetnames)
    workbook_ok = sheets == ["Sheet1", "Template", "Template Copy"]
    answer_ok = (
        parsed is not None
        and str(parsed.get("workbook")) == "tier_b_sheet_maintenance.xlsx"
        and sorted(str(item) for item in parsed.get("sheets", [])) == sheets
    )
    return answer_ok, workbook_ok, f"sheets={sheets}"


def _evaluate_table_setup(parsed: dict[str, Any] | None, workdir: Path) -> tuple[bool, bool, str]:
    workbook = workdir / "tier_b_table_setup.xlsx"
    ws = load_workbook(workbook)["Orders"]
    table_names = sorted(ws.tables.keys())
    header = [ws["A1"].value, ws["B1"].value, ws["C1"].value]
    workbook_ok = "OrdersTable" in table_names and header == ["Order", "Amount", "Status"]
    answer_ok = (
        parsed is not None
        and str(parsed.get("workbook")) == "tier_b_table_setup.xlsx"
        and str(parsed.get("sheet")) == "Orders"
        and str(parsed.get("table_name")) == "OrdersTable"
        and [str(item) for item in parsed.get("header", [])] == ["Order", "Amount", "Status"]
    )
    return answer_ok, workbook_ok, f"tables={table_names}"


def _evaluate_readback_verification(parsed: dict[str, Any] | None, workdir: Path) -> tuple[bool, bool, str]:
    workbook = workdir / "tier_b_readback.xlsx"
    ws = load_workbook(workbook)["Audit"]
    used_range = f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"
    workbook_ok = used_range == "A1:C4"
    answer_ok = (
        parsed is not None
        and str(parsed.get("workbook")) == "tier_b_readback.xlsx"
        and str(parsed.get("sheet")) == "Audit"
        and str(parsed.get("range")) == "A1:C4"
        and str(parsed.get("used_range")) == "A1:C4"
    )
    return answer_ok, workbook_ok, f"used_range={used_range}"


def _evaluate_row_maintenance(parsed: dict[str, Any] | None, workdir: Path) -> tuple[bool, bool, str]:
    workbook = workdir / "tier_b_row_ops.xlsx"
    ws = load_workbook(workbook)["Backlog"]
    rows = [ws["A2"].value, ws["A3"].value, ws["A4"].value]
    statuses = [ws["B2"].value, ws["B3"].value, ws["B4"].value]
    workbook_ok = rows == ["Task A", "Hotfix", "Task B"] and statuses == ["Open", "In Progress", "Done"]
    answer_ok = (
        parsed is not None
        and str(parsed.get("workbook")) == "tier_b_row_ops.xlsx"
        and str(parsed.get("sheet")) == "Backlog"
        and [str(item) for item in parsed.get("rows", [])] == ["Task A", "Hotfix", "Task B"]
    )
    return answer_ok, workbook_ok, f"rows={rows}"


def _evaluate_column_maintenance(parsed: dict[str, Any] | None, workdir: Path) -> tuple[bool, bool, str]:
    workbook = workdir / "tier_b_column_ops.xlsx"
    ws = load_workbook(workbook)["Metrics"]
    header = [ws["A1"].value, ws["B1"].value, ws["C1"].value]
    row2 = [ws["A2"].value, ws["B2"].value, ws["C2"].value]
    workbook_ok = header == ["Metric", "Target", "Actual"] and row2 == ["Latency", 95, 91]
    answer_ok = (
        parsed is not None
        and str(parsed.get("workbook")) == "tier_b_column_ops.xlsx"
        and str(parsed.get("sheet")) == "Metrics"
        and [str(item) for item in parsed.get("header", [])] == ["Metric", "Target", "Actual"]
    )
    return answer_ok, workbook_ok, f"header={header} row2={row2}"


def _seed_formula_flow_workbook(workdir: Path) -> Path:
    workbook = workdir / "tier_b_formula_flow.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws.append(["Item", "Q1", "Q2", "Total"])
    ws.append(["Compute", 12, 18, None])
    ws.append(["Storage", 10, 15, None])
    wb.save(workbook)
    return workbook


def _seed_layout_flow_workbook(workdir: Path) -> Path:
    workbook = workdir / "tier_b_layout_flow.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    ws["A1"] = "Operations Dashboard"
    ws["A3"] = "Tickets"
    ws["B3"] = 42
    ws["A4"] = "Latency"
    ws["B4"] = 91
    wb.save(workbook)
    return workbook


def _seed_sheet_maintenance_workbook(workdir: Path) -> Path:
    workbook = workdir / "tier_b_sheet_maintenance.xlsx"
    wb = Workbook()
    wb.active.title = "Sheet1"
    wb.create_sheet("Draft")
    wb.save(workbook)
    return workbook


def _seed_table_setup_workbook(workdir: Path) -> Path:
    workbook = workdir / "tier_b_table_setup.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    ws.append(["Order", "Amount", "Status"])
    ws.append(["A-100", 1200, "Open"])
    ws.append(["A-101", 980, "Closed"])
    ws.append(["A-102", 640, "Open"])
    wb.save(workbook)
    return workbook


def _seed_readback_verification_workbook(workdir: Path) -> Path:
    workbook = workdir / "tier_b_readback.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Audit"
    ws.append(["Check", "Owner", "Status"])
    ws.append(["Backups", "Ops", "Pass"])
    ws.append(["Alerts", "SRE", "Pass"])
    ws.append(["Access", "IT", "Review"])
    wb.save(workbook)
    return workbook


def _seed_row_maintenance_workbook(workdir: Path) -> Path:
    workbook = workdir / "tier_b_row_ops.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Backlog"
    ws.append(["Task", "Status"])
    ws.append(["Task A", "Open"])
    ws.append(["Task B", "Done"])
    ws.append(["Task C", "Blocked"])
    wb.save(workbook)
    return workbook


def _seed_column_maintenance_workbook(workdir: Path) -> Path:
    workbook = workdir / "tier_b_column_ops.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Metrics"
    ws.append(["Metric", "Actual", "Owner"])
    ws.append(["Latency", 91, "API"])
    ws.append(["Success", 88, "Infra"])
    wb.save(workbook)
    return workbook
